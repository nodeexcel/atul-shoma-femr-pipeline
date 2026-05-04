"""
FEMR Funds transformation service.

Reads the 'FEMR Funds' sheet from an uploaded workbook and reshapes it into
a long-format Output with one row per (Sequence, Quarter, Type).

Four types per (Sequence, Quarter):
  Committed      – Award Amount column
  Obligated      – Obligated column
  Expended       – Cash Collected column
  Remaining Cash – Obligated minus Expended

Quarters are discovered dynamically from row 4 of the FEMR Funds sheet.
Cells with 'QE M/D/YY' labels are quarter end dates; annual total columns
('FYxx Total', 'Total') are skipped automatically.

Output format: 'excel' returns xlsx bytes with Output tab only (no input tabs).
               'csv' returns UTF-8 CSV bytes.
"""
import csv
import io
import logging
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

logger = logging.getLogger('transform.services')

_OUTPUT_TYPES = [
    ('Committed',  0),
    ('Obligated',  1),
    ('Expended',   2),
]


def _discover_quarters(ws_femr) -> list[tuple]:
    """
    Dynamically build the quarter list by reading row 4 of the FEMR Funds sheet.
    Cells with a 'QE M/D/YY' label are quarter end dates; all other labels
    (e.g. 'FYxx Total', 'Total') are skipped automatically.

    Returns: list of (quarter_end_date, col_committed, col_obligated, col_expended)
             Columns are 1-indexed.
    """
    quarters = []
    for col in range(1, ws_femr.max_column + 1):
        raw = ws_femr.cell(row=4, column=col).value
        if raw is None:
            continue
        label = str(raw).strip()
        if not label.upper().startswith('QE '):
            continue
        date_str = label[3:].strip()
        try:
            qdate = datetime.strptime(date_str, '%m/%d/%y')
        except ValueError:
            continue
        quarters.append((qdate, col, col + 1, col + 2))
    return quarters


def safe_float(val) -> float:
    """Convert a cell value to float, returning 0.0 for None or non-numeric."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _read_sequences(ws_femr, ws_output=None) -> list[str]:
    """
    Return the ordered sequence list.
    Always derives sequences from FEMR Funds (source of truth).
    If an Output template exists, its sequence order is used to sort the list
    so output order stays consistent across runs. New sequences not in the
    template are appended at the end.
    """
    sequences = []
    seen = set()
    for row in ws_femr.iter_rows(min_row=7, max_row=306, values_only=True):
        raw = row[3]  # Column D
        if raw is None:
            continue
        seq = str(raw).strip()
        if seq and seq not in seen:
            sequences.append(seq)
            seen.add(seq)

    if ws_output is not None:
        template_order = {}
        for row in ws_output.iter_rows(min_row=2, max_row=135, values_only=True):
            if row[0] is not None:
                seq = str(row[0]).strip()
                if seq and seq not in template_order:
                    template_order[seq] = len(template_order)
        if template_order:
            sequences.sort(key=lambda s: template_order.get(s, len(template_order)))

    return sequences


def _aggregate_data(ws_femr, sequences: list[str], quarters: list[tuple]) -> dict:
    """
    Aggregate (sum) quarterly financials per sequence from FEMR Funds sheet.
    Returns: { sequence -> { quarter_date -> [committed, obligated, expended] } }
    Sequences appearing on multiple rows are summed together.
    """
    seq_set = set(sequences)
    data = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0, 0.0]))

    for row in ws_femr.iter_rows(min_row=7, max_row=306, values_only=True):
        raw = row[3]
        if raw is None:
            continue
        seq = str(raw).strip()
        if seq not in seq_set:
            continue
        for (qdate, cc, co, ce) in quarters:
            data[seq][qdate][0] += safe_float(row[cc - 1])
            data[seq][qdate][1] += safe_float(row[co - 1])
            data[seq][qdate][2] += safe_float(row[ce - 1])

    return data


def _build_output_rows(sequences: list[str], data: dict, quarters: list[tuple]) -> list[tuple]:
    """
    Build the flat list of output rows.
    Structure: for each quarter → Committed block, Obligated block,
               Expended block, Remaining Cash block, each with all sequences.
    """
    rows = []
    for (qdate, *_) in quarters:
        for type_name, idx in _OUTPUT_TYPES:
            for seq in sequences:
                rows.append((seq, qdate, type_name, data[seq][qdate][idx]))
        for seq in sequences:
            remaining = data[seq][qdate][1] - data[seq][qdate][2]
            rows.append((seq, qdate, 'Remaining Cash', remaining))
    return rows



def _to_excel_bytes(output_rows: list[tuple]) -> bytes:
    """Build a fresh workbook with only the Output tab and return xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Output'

    bold = Font(bold=True)
    for col, header in enumerate(['Sequence', 'Qtr Date', 'Type', 'Amount'], 1):
        ws.cell(row=1, column=col, value=header).font = bold

    for i, (seq, qdate, type_name, amount) in enumerate(output_rows, start=2):
        ws.cell(row=i, column=1, value=seq)
        date_cell = ws.cell(row=i, column=2, value=qdate)
        date_cell.number_format = 'mm-dd-yy'
        ws.cell(row=i, column=3, value=type_name)
        ws.cell(row=i, column=4, value=amount)

    last_row = 1 + len(output_rows)
    ws.auto_filter.ref = f'A1:D{last_row}'
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _to_csv_bytes(output_rows: list[tuple]) -> bytes:
    """Serialise output rows to UTF-8 CSV bytes."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Sequence', 'Qtr Date', 'Type', 'Amount'])
    for seq, qdate, type_name, amount in output_rows:
        writer.writerow([seq, qdate.strftime('%m/%d/%Y'), type_name, round(amount, 2)])
    return buf.getvalue().encode('utf-8')


def run_transform(input_path: str, fmt: str = 'excel') -> bytes:
    """
    Read the FEMR Funds workbook at *input_path*, build the output rows,
    and return the result as bytes in the requested format ('excel' or 'csv').
    Excel output contains only the Output tab — no input tabs carried over.
    """
    logger.info("Transform started: %s (format=%s)", input_path, fmt)

    wb_values = load_workbook(input_path, data_only=True)
    ws_femr = wb_values['FEMR Funds']
    ws_output_template = wb_values['Output'] if 'Output' in wb_values.sheetnames else None

    quarters = _discover_quarters(ws_femr)
    sequences = _read_sequences(ws_femr, ws_output_template)
    data = _aggregate_data(ws_femr, sequences, quarters)
    output_rows = _build_output_rows(sequences, data, quarters)

    logger.info(
        "Transform complete: %d sequences × %d quarters × 4 types = %d rows",
        len(sequences), len(quarters), len(output_rows),
    )

    if fmt == 'csv':
        return _to_csv_bytes(output_rows)
    return _to_excel_bytes(output_rows)
