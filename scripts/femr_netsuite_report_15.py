"""
FEMR NetSuite Report Generator (v15)
=====================================
Matches the new FEMR Export Template 041526.xlsx layout.

Changes from v14:
  - Chart quarter labels pushed outside/below the plot area. Previously
    tickLblPos="nextTo" placed labels on the zero line inside the grid,
    overlapping with chart content. Changed to tickLblPos="low" so labels
    always render below the plot area regardless of Y-axis range.
    Josh confirmed this during 2026-04-28 deployment meeting.

Changes from v12:
  - Chart legend moved to RIGHT side (was bottom). Bottom legend overlapped
    X-axis quarter labels. Josh confirmed right side is correct (2026-04-27).
  - Chart X-axis stays at bottom when Y values are negative. Previously
    catAx used crosses="autoZero", which placed the category axis at Y=0 —
    when all values were negative this pushed quarter labels to the top.
    Fixed by setting crosses="min" in _patch_chart_axes() so the X-axis
    always anchors at the minimum Y value (bottom of chart).

Changes from v11:
  - Chart axis labels: Y-axis (dollar values) and X-axis (quarter labels) now
    visible. Root cause: chart.y_axis.delete and chart.x_axis.delete were not
    set to False, hiding axis tick labels. Y-axis numFmt set to '$#,##0.00'.
  - Pre-FY2020 cumulative starting balance: fetches FY2016-2019 historical data
    and seeds Q1 FY2020 cumulative opening balance. Previously started at 0.
    Adds ~80 extra API calls per sequence (4 years × 4 quarters × 5 accounts).

Changes from v10:
  - File splitting: groups with >SPLIT_SIZE sequences are split into multiple
    files of 50 tabs each (e.g. ADP 247 tabs → 5 files: 001-050, 051-100,
    101-150, 151-200, 201-247). Groups under 50 stay as one file.
    Controlled by --split-size (default 50).

Changes from v5:
  - Bulk metadata preload: 2 MV queries at startup (~90s) replace per-sequence
    metadata calls (~13 min → ~90s total for full run)
  - Multi-rollup sequences automatically detected from bulk data
  - Per-sequence fallback kept for edge cases not covered by bulk filter

Changes from v4:
  - NEW 11-row header block (REPORTING GROUP, SEQUENCE, ROLLUP #, PROJECT NAME,
    TECHNICAL POINT OF CONTACT, PROJECT TYPE, SERVICE, COLOR OF MONEY, ACRN,
    AVAILABLE FUNDS)
  - Data shifted to rows 17+ (was row 11)
  - Cumulative section moved to rows 51-56 (was 48-54)
  - Removed Revised Plan row
  - Removed govt awards / cash collected rows
  - Added Remaining Cash inside Contracting section (Obligated − Expended)
  - Rollup now used as API parameter — one tab per (sequence, rollup) pair
  - Dynamic mapping loaded from docs/Re_ FEMR /GROUP MAPPING.xlsx
  - Dynamic quarter range: FY2020 to latest quarter with data
  - Line chart embedded below cumulative section
  - All metadata fetched from /mv_femr_report/ ORDS filter query

Script supports all reporting groups: ADP, Comml, Internal, OGA, WFD.

Layout per tab:
  Row 1     : NextFlex
  Row 2     : REPORTING GROUP - TYPE A | value (ADP, Comml, ...)
  Row 3     : SEQUENCE | sequence code
  Row 4     : ROLLUP # | rollup number
  Row 5     : PROJECT NAME | legal name
  Row 6     : TECHNICAL POINT OF CONTACT | name
  Row 7     : PROJECT TYPE | type_b (comma-separated possible)
  Row 8     : SERVICE | service
  Row 9     : COLOR OF MONEY | (blank)
  Row 10    : ACRN | acrn
  Row 11    : AVAILABLE FUNDS | amount
  Row 12-14 : (blank)
  Row 15    : FYE headers (merged across 4 Q cols each)
  Row 16    : Q1 | Q2 | Q3 | Q4 ... | (Grand Total)
  Row 17-29 : Actuals (Labor Hours + 12 accounts)
  Row 30    : ACTUALS Total
  Row 31-43 : Budgeted (Labor Hours + 12 accounts)
  Row 44    : BUDGETED Total
  Row 45-48 : Contracting (Committed, Obligated, Expended, Remaining Cash)
  Row 49-50 : (blank)
  Row 51    : Q labels (Q1 FY20 ...)
  Row 52    : Total Committed (cumulative)
  Row 53    : Total Obligated (cumulative)
  Row 54    : Total Expended (cumulative)
  Row 55    : Budgeted Plan (cumulative)
  Row 56    : Actual (cumulative)
  Row 59+   : LineChart (5 series)

Usage:
    python femr_netsuite_report_7.py --sequence 2ADP001        # single tab test
    python femr_netsuite_report_7.py --group ADP               # all ADP sequences
    python femr_netsuite_report_7.py                           # all 5 groups
    python femr_netsuite_report_7.py --workers 30              # tune concurrency
"""

import argparse
import logging
import time
import urllib.request
import urllib.parse
import json
import sys
import os
import io
import re
import zipfile as _zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── API ───────────────────────────────────────────────────────────────────────
BASE_URL = "https://g22673cc0c08b7a-oax2132513753.adb.us-ashburn-1.oraclecloudapps.com/ords/oax_user"
NETAMOUNT_URL = f"{BASE_URL}/femr/netamount/"
MV_URL = f"{BASE_URL}/mv_femr_report/"

# ── Fiscal scope ──────────────────────────────────────────────────────────────
PRE_FISCAL_YEAR_START = 2016   # earliest year for pre-FY2020 cumulative seeding
FISCAL_YEAR_START = 2020
FISCAL_YEAR_END_DEFAULT = 2026   # used if MV lookup for latest year fails
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

# ── Account definitions ───────────────────────────────────────────────────────
# (row_label, api_account_name)
# api_account_name=None → statistical account, always blank
ACTUALS_BUDGET_ACCOUNTS = [
    ("Labor Hours statistical account", None),
    ("Labor Cost 5001",                 "5001 DIR : Direct Labor"),
    ("Fringe 5990",                     "5990 ALLO : Allo Fringe"),
    ("Travel 5004",                     "5004 DIR : Direct Travel"),
    ("Subcontracting 5005",             "5005 DIR : Subrecipient Costs"),
    ("Consulting 5002",                 "5002 DIR : Direct Consulting"),
    ("Equipment 5010",                  "5010 DIR : EQ & Materials (NO OH)"),
    ("Equipment 5008",                  "5008 DIR : Direct Equipment"),
    ("Other Direct Costs 5009",         "5009 DIR : Direct Other Costs"),
    ("Material 5003",                   "5003 DIR : Direct Materials"),
    ("Sub K Overhead 5992",             "5992 ALLO : Allo SubK OH"),
    ("Sub K Overhead 5993",             "5993 ALLO : DNU ALLO G and A OH WFD"),
    ("G&A 5991",                        "5991 ALLO : Allo G and A"),
]

CONTRACTING_ACCOUNTS = [
    ("Committed",  "Committed"),
    ("Obligated",  "Obligated"),
    ("Expended",   "Expended"),
]

# ── File splitting ────────────────────────────────────────────────────────────
SPLIT_SIZE = 50   # max tabs per output file; groups under this stay as one file

# ── Layout constants ─────────────────────────────────────────────────────────
# Header block
ROW_FYE      = 15
ROW_Q        = 16
# Actuals
ROW_ACTUALS_START = 17   # Labor Hours row
ROW_ACTUALS_END   = 29   # G&A row
ROW_ACTUALS_TOTAL = 30
# Budgeted
ROW_BUDGET_START = 31
ROW_BUDGET_END   = 43
ROW_BUDGET_TOTAL = 44
# Contracting
ROW_COMMITTED      = 45
ROW_OBLIGATED      = 46
ROW_EXPENDED       = 47
ROW_REMAINING_CASH = 48
# Cumulative
ROW_CUM_Q_LABELS   = 51
ROW_CUM_COMMITTED  = 52
ROW_CUM_OBLIGATED  = 53
ROW_CUM_EXPENDED   = 54
ROW_CUM_BUDGETED   = 55
ROW_CUM_ACTUAL     = 56
# Chart
ROW_CHART_START    = 59

COL_DATA_START = 3  # col C

# ── Styles ────────────────────────────────────────────────────────────────────
FYE_GREEN  = PatternFill("solid", fgColor="C1FFB0")
FYE_BLUE   = PatternFill("solid", fgColor="A3D1FF")
GREY_FILL  = PatternFill("solid", fgColor="F5F4F2")

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
VCENTER_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

NUM_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# ─────────────────────────────────────────────────────────────────────────────
# DYNAMIC QUARTER DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def detect_latest_fiscal_quarter(max_year: int = 2035) -> tuple:
    """
    Find the latest (year, quarter) that has any data in the MV.
    Probes from most recent year/quarter backward using limit=1 MV queries.
    Each probe is fast (~1s). Worst case: ~60 probes if no data found until 2020.
    """
    logger.info("Detecting latest fiscal quarter with data...")
    for year in range(max_year, FISCAL_YEAR_START - 1, -1):
        fye = f"FYE 9/30/{year}"
        for q in reversed(QUARTERS):  # Q4 → Q3 → Q2 → Q1
            q_filter = json.dumps({"fiscal_year_end": fye, "fiscal_quarter": q})
            url = f"{MV_URL}?q={urllib.parse.quote(q_filter)}&limit=1"
            try:
                data = _http_get(url)
                if data.get("items"):
                    logger.info("Latest quarter with data: %s %s", q, fye)
                    return year, q
            except Exception as exc:
                logger.warning("Quarter probe failed %s %s: %s", fye, q, exc)
    logger.warning("Could not detect latest quarter — falling back to Q4 FY%d",
                   FISCAL_YEAR_END_DEFAULT)
    return FISCAL_YEAR_END_DEFAULT, "Q4"


# ─────────────────────────────────────────────────────────────────────────────
# MAPPING FILE LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_sequence_registry(mapping_path: str) -> list:
    """
    Load the GROUP MAPPING file and return a list of unique sequences with
    their reporting group. Returns: list of dicts with keys:
        sequence, reporting_group, type_a_category
    """
    wb = load_workbook(mapping_path, data_only=True, read_only=True)
    ws = wb["SEQUENCE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # Dedup on sequence — take the first occurrence
    seen = {}
    for row in rows:
        if not row or row[3] is None:
            continue
        rg, category, type_b, seq = row[0], row[1], row[2], row[3]
        if seq not in seen:
            seen[seq] = {
                "sequence": seq,
                "reporting_group": rg,
                "type_a_category": category,
            }
    return list(seen.values())


# ─────────────────────────────────────────────────────────────────────────────
# API FETCHING
# ─────────────────────────────────────────────────────────────────────────────

def _http_get(url: str, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                return json.loads(r.read())
        except Exception as exc:
            if attempt == retries - 1:
                raise
            time.sleep(1.5 ** attempt)
    return {}


def _identifier_for(row: dict) -> tuple:
    """
    Return the unique tab identifier for an MV row.

    Rules:
      - For Orphans (class='Orphan'): use project_number as the identifier.
      - For Rollups (Parent/Child): use display_rollup_num. Parent + Children
        all share the same rollup_num, so they aggregate into one tab.

    Returns: (identifier_type, identifier_value)
        identifier_type: 'rollup' or 'orphan'
        identifier_value: the rollup_num or project_number
    """
    cls = (row.get("class") or "").strip()
    if cls == "Orphan":
        return ("orphan", row.get("project_number"))
    # Parent or Child → aggregate into rollup
    return ("rollup", row.get("display_rollup_num"))


def _row_to_meta(i: dict) -> dict:
    """
    Extract metadata fields from an MV row.

    Key changes from v6:
      - SERVICE: use display_type_c (comma-separated full list), fall back to service
      - ROLLUP: for orphans, use project_number; for rollups, use display_rollup_num
    """
    cls = (i.get("class") or "").strip()
    is_orphan = cls == "Orphan"

    rollup_display = i.get("display_rollup_num")
    if is_orphan and not rollup_display:
        rollup_display = i.get("project_number")

    return {
        "sequence":               i.get("display_sequence"),
        "rollup":                 rollup_display,
        "rollup_num":             i.get("display_rollup_num"),   # for API calls (may be None)
        "project_number":         i.get("project_number"),
        "class":                  cls,
        "is_orphan":              is_orphan,
        "project_legal_name":     i.get("display_project_legalname"),
        "technical_point_of_contact": i.get("technical_point_of_contact"),
        "acrn":                   i.get("display_acrn"),
        "available_funds":        i.get("available_funds") or 0,
        "type_b":                 i.get("display_type_b") or "",
        # SERVICE: prefer display_type_c (full comma-separated) over service
        "service":                i.get("display_type_c") or i.get("service") or "",
        "reporting_group":        i.get("reporting_group_type_a"),
    }


def fetch_sequence_identifiers(sequence: str) -> list:
    """
    Query MV for one sequence and return all distinct (tab identifier) metadata.

    For sequences like CC007 that have both Rollup 4006 (parent+children) AND
    an Orphan (project 960011), returns 2 distinct metadata dicts — one per tab.

    Paginates until all rows are fetched (most sequences: 1-2 pages).
    """
    all_ids = {}   # (id_type, id_value) -> first metadata dict
    offset = 0
    page = 0
    while True:
        q = json.dumps({"display_sequence": sequence})
        url = f"{MV_URL}?q={urllib.parse.quote(q)}&limit=2000&offset={offset}"
        try:
            data = _http_get(url)
        except Exception as exc:
            logger.warning("Metadata fetch failed for %s: %s", sequence, exc)
            break
        items = data.get("items", [])
        page += 1
        for i in items:
            id_tuple = _identifier_for(i)
            if id_tuple[1] is None:
                continue
            if id_tuple not in all_ids:
                all_ids[id_tuple] = _row_to_meta(i)
        if not data.get("hasMore"):
            break
        offset += len(items)
        if page > 5:
            logger.warning("%s: metadata hit 5-page safety cap", sequence)
            break

    return list(all_ids.values())


def preload_all_identifiers(sequences: list, workers: int = 20) -> dict:
    """
    Parallel per-sequence metadata fetch to find ALL distinct identifiers
    (rollups + orphans) for every sequence.

    Strategy: Query MV by display_sequence for each sequence, in parallel.
    Each query is fast (~1-3s per sequence). This replaces v6's bulk preload
    and catches orphans the bulk filter missed.

    Returns: {sequence: [metadata_dict, ...]} — list because a sequence
    may have multiple identifiers (e.g. CC007 has rollup 4006 + orphan 960011).
    """
    logger.info("Fetching metadata for %d sequences (parallel, %d workers)...",
                len(sequences), workers)
    t0 = time.time()

    result = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(fetch_sequence_identifiers, seq): seq
            for seq in sequences
        }
        done = 0
        for future in as_completed(future_map):
            seq = future_map[future]
            metas = future.result()
            result[seq] = metas
            done += 1
            if done % 50 == 0:
                logger.info("  metadata progress: %d/%d", done, len(sequences))

    total_tabs = sum(len(v) for v in result.values())
    logger.info("Preloaded metadata for %d sequences (%d total tabs) in %.1fs",
                len(result), total_tabs, time.time() - t0)
    return result


def _fetch_netamount(sequence: str, fye: str, quarter: str,
                     segment: str, account_name: str, retries: int = 3) -> float:
    """
    Fast: use /femr/netamount/ API. Returns combined sum across ALL
    identifiers for the sequence (cannot differentiate rollup vs orphan).
    Use only when sequence has exactly one identifier.
    """
    params = {
        "display_sequence": sequence,
        "fiscal_year_end":  fye,
        "fiscal_quarter":   quarter,
        "segment":          segment,
        "account_name":     account_name,
    }
    url = f"{NETAMOUNT_URL}?{urllib.parse.urlencode(params)}"
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                data = json.loads(r.read())
            val = data.get("items", [{}])[0].get("total_netamount")
            return float(val) if val is not None else 0.0
        except Exception as exc:
            if attempt == retries - 1:
                logger.warning("NETAMOUNT FAILED %s %s %s %s: %s",
                               sequence, fye, quarter, account_name, exc)
                return 0.0
            time.sleep(1.5 ** attempt)
    return 0.0


def _fetch_mv_by_identifier(sequence: str, fye: str, quarter: str,
                             segment: str, account_name: str,
                             retries: int = 3) -> dict:
    """
    Slow: use /mv_femr_report/ directly with tight filter, sum netamount
    client-side per identifier (rollup OR orphan project_number).

    Returns: {(id_type, id_value): total_netamount}
    """
    q = json.dumps({
        "display_sequence": sequence,
        "fiscal_year_end":  fye,
        "fiscal_quarter":   quarter,
        "segment":          segment,
        "account_name":     account_name,
    })
    url = f"{MV_URL}?q={urllib.parse.quote(q)}&limit=500"
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                data = json.loads(r.read())
            sums = {}
            for i in data.get("items", []):
                id_tuple = _identifier_for(i)
                if id_tuple[1] is None:
                    continue
                sums.setdefault(id_tuple, 0.0)
                sums[id_tuple] += i.get("netamount") or 0.0
            return sums
        except Exception as exc:
            if attempt == retries - 1:
                logger.warning("MV FAILED %s %s %s %s: %s",
                               sequence, fye, quarter, account_name, exc)
                return {}
            time.sleep(1.5 ** attempt)
    return {}


def _build_tasks(fiscal_years: list, latest_quarter: str = "Q4") -> list:
    """Build (fye, quarter, segment, account) tuples up to latest_quarter in the last year."""
    last_year = fiscal_years[-1]
    last_q_idx = QUARTERS.index(latest_quarter)
    tasks = []
    for year in fiscal_years:
        fye = f"FYE 9/30/{year}"
        quarters = QUARTERS[:last_q_idx + 1] if year == last_year else QUARTERS
        for quarter in quarters:
            for _, account in ACTUALS_BUDGET_ACCOUNTS:
                if account is not None:
                    tasks.append((fye, quarter, "ACTUALS",  account))
                    tasks.append((fye, quarter, "BUDGETED", account))
            for _, account in CONTRACTING_ACCOUNTS:
                tasks.append((fye, quarter, "CONTRACTING", account))
    return tasks


def _sum_pre_fy2020(raw: dict) -> dict:
    """
    Given raw {(fye, q, seg, account): amount} for FY2016-2019,
    return pre-FY2020 cumulative sums for each cumulative metric.
    """
    pre_years = list(range(PRE_FISCAL_YEAR_START, FISCAL_YEAR_START))
    fyes = {f"FYE 9/30/{y}" for y in pre_years}

    committed = sum(v for (fye, q, seg, acc), v in raw.items()
                    if fye in fyes and seg == "CONTRACTING" and acc == "Committed")
    obligated = sum(v for (fye, q, seg, acc), v in raw.items()
                    if fye in fyes and seg == "CONTRACTING" and acc == "Obligated")
    expended  = sum(v for (fye, q, seg, acc), v in raw.items()
                    if fye in fyes and seg == "CONTRACTING" and acc == "Expended")
    budgeted  = sum(v for (fye, q, seg, acc), v in raw.items()
                    if fye in fyes and seg == "BUDGETED")
    actual    = sum(v for (fye, q, seg, acc), v in raw.items()
                    if fye in fyes and seg == "ACTUALS")
    return {"committed": committed, "obligated": obligated, "expended": expended,
            "budgeted": budgeted, "actual": actual}


def fetch_pre_fy2020_single(sequence: str, workers: int = 20) -> dict:
    """Fetch FY2016-2019 data for a single-identifier sequence via netamount API."""
    pre_years = list(range(PRE_FISCAL_YEAR_START, FISCAL_YEAR_START))
    tasks = _build_tasks(pre_years, latest_quarter="Q4")
    results = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_netamount, sequence, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            results[future_map[future]] = future.result()
    return _sum_pre_fy2020(results)


def fetch_pre_fy2020_multi(sequence: str, workers: int = 20) -> dict:
    """Fetch FY2016-2019 data for a multi-identifier sequence via MV. Returns per-id sums."""
    pre_years = list(range(PRE_FISCAL_YEAR_START, FISCAL_YEAR_START))
    tasks = _build_tasks(pre_years, latest_quarter="Q4")
    per_identifier = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_mv_by_identifier, sequence, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            key = future_map[future]
            for id_tuple, amount in future.result().items():
                per_identifier.setdefault(id_tuple, {})[key] = amount
    return {id_tuple: _sum_pre_fy2020(raw) for id_tuple, raw in per_identifier.items()}


def fetch_financials_single(sequence: str, fiscal_years: list,
                              workers: int = 20, latest_quarter: str = "Q4") -> dict:
    """
    For sequences with a single identifier: use fast netamount API.
    Returns {(fye, q, seg, account): amount}
    """
    tasks = _build_tasks(fiscal_years, latest_quarter)
    results = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_netamount, sequence, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            key = future_map[future]
            results[key] = future.result()
    return results


def fetch_financials_multi(sequence: str, fiscal_years: list,
                            workers: int = 20, latest_quarter: str = "Q4") -> dict:
    """
    For sequences with multiple identifiers: query MV per cell,
    split by identifier.

    Returns: {id_tuple: {(fye, q, seg, account): amount}}
    """
    tasks = _build_tasks(fiscal_years, latest_quarter)
    per_identifier = {}   # id_tuple -> {task_key -> amount}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_mv_by_identifier, sequence, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            task_key = future_map[future]
            sums_per_id = future.result()
            for id_tuple, amount in sums_per_id.items():
                per_identifier.setdefault(id_tuple, {})
                per_identifier[id_tuple][task_key] = amount
    return per_identifier


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────

def _fye(year: int) -> str:
    return f"FYE 9/30/{year}"


def _q_label(year: int, q: str) -> str:
    # Non-breaking space ( ) prevents Excel multi-level axis grouping
    # while appearing identical to a regular space visually
    return f"{q} FY{str(year)[2:]}"


def _write_header_block(ws, meta: dict):
    """Write rows 1-11 header block."""
    fields = [
        ("NextFlex",                      ""),
        ("REPORTING GROUP - TYPE A",      meta.get("reporting_group") or ""),
        ("SEQUENCE",                      meta.get("sequence") or ""),
        ("ROLLUP #",                      meta.get("rollup") or ""),
        ("PROJECT NAME",                  meta.get("project_legal_name") or ""),
        ("TECHNICAL POINT OF CONTACT",    meta.get("technical_point_of_contact") or ""),
        ("PROJECT TYPE",                  meta.get("type_b") or ""),
        ("SERVICE",                       meta.get("service") or ""),
        ("COLOR OF MONEY",                ""),
        ("ACRN",                          meta.get("acrn") or ""),
        ("AVAILABLE FUNDS",               meta.get("available_funds") or 0),
    ]
    for r, (label, value) in enumerate(fields, start=1):
        la = ws.cell(row=r, column=1, value=label)
        la.font = BOLD
        vc = ws.cell(row=r, column=2, value=value)
        if label == "AVAILABLE FUNDS" and isinstance(value, (int, float)):
            vc.number_format = NUM_FMT
            vc.alignment = RIGHT


def _write_col_headers(ws, fiscal_years: list, latest_quarter: str = "Q4") -> dict:
    """Write FYE (row 15) and Q (row 16) headers, return col_map.
    Stops at latest_quarter for the last fiscal year."""
    col = COL_DATA_START
    col_map = {}
    fills = [FYE_GREEN, FYE_BLUE]
    last_year = fiscal_years[-1]
    last_q_idx = QUARTERS.index(latest_quarter)

    for idx, year in enumerate(fiscal_years):
        fye = _fye(year)
        fill = fills[idx % 2]
        quarters = QUARTERS[:last_q_idx + 1] if year == last_year else QUARTERS
        n = len(quarters)
        ws.merge_cells(start_row=ROW_FYE, start_column=col,
                       end_row=ROW_FYE,   end_column=col + n - 1)
        c = ws.cell(row=ROW_FYE, column=col, value=fye)
        c.alignment = CENTER
        c.font = BOLD
        c.fill = fill
        for qi, q in enumerate(quarters):
            ws.cell(row=ROW_FYE, column=col + qi).fill = fill
            qc = ws.cell(row=ROW_Q, column=col + qi, value=q)
            qc.alignment = CENTER
            qc.font = BOLD
            qc.fill = fill
            col_map[(fye, q)] = col + qi
        col += n

    # Grand Total column
    gt = ws.cell(row=ROW_FYE, column=col, value="Grand Total")
    gt.alignment = CENTER
    gt.font = BOLD
    gt.fill = GREY_FILL
    ws.cell(row=ROW_Q, column=col).fill = GREY_FILL
    col_map["total"] = col

    return col_map


def _write_data_row(ws, row: int, data: dict, fye_q_pairs: list,
                    col_map: dict, segment: str, api_account: Optional[str]):
    """Write a single data row with values + Grand Total SUM formula."""
    first_col = None
    last_col = None
    for (fye, q) in fye_q_pairs:
        col = col_map.get((fye, q))
        if col is None:
            continue
        if first_col is None:
            first_col = col
        last_col = col
        amount = 0.0 if api_account is None else data.get((fye, q, segment, api_account), 0.0)
        if amount != 0.0:
            c = ws.cell(row=row, column=col, value=amount)
            c.number_format = NUM_FMT
            c.alignment = RIGHT

    # Grand Total
    total_col = col_map.get("total")
    if total_col and first_col and last_col:
        fl = get_column_letter(first_col)
        ll = get_column_letter(last_col)
        tc = ws.cell(row=row, column=total_col,
                     value=f"=SUM({fl}{row}:{ll}{row})")
        tc.number_format = NUM_FMT
        tc.alignment = RIGHT
        tc.fill = GREY_FILL


def _write_section_total_row(ws, row: int, data_start_row: int, col_map: dict):
    """Write a SUM total row across account rows (skips Labor Hours row)."""
    sum_start = data_start_row + 1  # skip Labor Hours
    for col_idx in col_map.values():
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=row, column=col_idx,
                    value=f"=SUM({col_letter}{sum_start}:{col_letter}{row - 1})")
        c.number_format = NUM_FMT
        c.alignment = RIGHT
        c.fill = GREY_FILL


def _write_remaining_cash_row(ws, col_map: dict):
    """Remaining Cash = Obligated - Expended per quarter."""
    total_col = col_map.get("total")
    first_col = None
    last_col = None
    for key, col in col_map.items():
        if not isinstance(key, tuple):
            continue
        if first_col is None or col < first_col:
            first_col = col
        if last_col is None or col > last_col:
            last_col = col
        cl = get_column_letter(col)
        c = ws.cell(row=ROW_REMAINING_CASH, column=col,
                    value=f"={cl}{ROW_OBLIGATED}-{cl}{ROW_EXPENDED}")
        c.number_format = NUM_FMT
        c.alignment = RIGHT

    if total_col and first_col and last_col:
        fl = get_column_letter(first_col)
        ll = get_column_letter(last_col)
        tc = ws.cell(row=ROW_REMAINING_CASH, column=total_col,
                     value=f"=SUM({fl}{ROW_REMAINING_CASH}:{ll}{ROW_REMAINING_CASH})")
        tc.number_format = NUM_FMT
        tc.alignment = RIGHT
        tc.fill = GREY_FILL


def _write_cumulative_section(ws, fiscal_years: list, col_map: dict,
                               latest_quarter: str = "Q4",
                               pre_cumulative: Optional[dict] = None):
    """
    Rows 51-56: cumulative running totals.
    Row 51: Q labels starting at col B
    Row 52: Total Committed (cumulative from row 45)
    Row 53: Total Obligated (cumulative from row 46)
    Row 54: Total Expended (cumulative from row 47)
    Row 55: Budgeted Plan (cumulative from row 44)
    Row 56: Actual (cumulative from row 30)

    pre_cumulative: {committed, obligated, expended, budgeted, actual} sums from
    FY2016-2019 — seeded into Q1 FY2020 opening balance formula.
    """
    pre = pre_cumulative or {}
    last_year = fiscal_years[-1]
    last_q_idx = QUARTERS.index(latest_quarter)

    # Row 51: Q labels - start at col B (col 2) per new template
    col_b = 2
    for year in fiscal_years:
        quarters = QUARTERS[:last_q_idx + 1] if year == last_year else QUARTERS
        for q in quarters:
            c = ws.cell(row=ROW_CUM_Q_LABELS, column=col_b, value=_q_label(year, q))
            c.alignment = CENTER
            c.font = BOLD
            col_b += 1

    formula_defs = [
        (ROW_CUM_COMMITTED, "Total Committed",  ROW_COMMITTED,     "committed"),
        (ROW_CUM_OBLIGATED, "Total Obligated",  ROW_OBLIGATED,     "obligated"),
        (ROW_CUM_EXPENDED,  "Total Expended",   ROW_EXPENDED,      "expended"),
        (ROW_CUM_BUDGETED,  "Budgeted Plan",    ROW_BUDGET_TOTAL,  "budgeted"),
        (ROW_CUM_ACTUAL,    "Actual",           ROW_ACTUALS_TOTAL, "actual"),
    ]

    for (frow, label, src_row, pre_key) in formula_defs:
        ws.cell(row=frow, column=1, value=label).font = BOLD

        pre_val = pre.get(pre_key, 0.0) or 0.0

        # Column B holds first quarter; each subsequent adds to prev cumulative
        prev_col_letter = None
        col_b = 2
        for year in fiscal_years:
            quarters = QUARTERS[:last_q_idx + 1] if year == last_year else QUARTERS
            for q in quarters:
                src_col = col_map.get((_fye(year), q))
                if src_col is None:
                    col_b += 1
                    continue
                src_col_letter = get_column_letter(src_col)
                this_col_letter = get_column_letter(col_b)

                if prev_col_letter is None:
                    # Q1 FY2020: seed with pre-FY2020 historical total if non-zero
                    if pre_val:
                        formula = f"={pre_val}+{src_col_letter}{src_row}"
                    else:
                        formula = f"={src_col_letter}{src_row}"
                else:
                    formula = f"={src_col_letter}{src_row}+{prev_col_letter}{frow}"

                c = ws.cell(row=frow, column=col_b, value=formula)
                c.number_format = NUM_FMT
                c.alignment = RIGHT
                prev_col_letter = this_col_letter
                col_b += 1


def _patch_chart_axes(xlsx_path: str) -> None:
    """Post-process saved xlsx: replace catAx to exactly match client template structure.

    openpyxl generates incomplete catAx XML (wrong axPos, missing tickLblPos, auto,
    crosses, lblAlgn, rotation). Full replacement ensures Excel Online renders
    diagonal single-row labels matching the FEMR Export Template.
    """
    raw = Path(xlsx_path).read_bytes()
    in_buf = io.BytesIO(raw)
    out_buf = io.BytesIO()

    _catax_pat = re.compile(r'<catAx>.*?</catAx>', re.DOTALL)

    with _zipfile.ZipFile(in_buf, 'r') as zin, \
         _zipfile.ZipFile(out_buf, 'w', compression=_zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if re.match(r'xl/charts/chart\d+\.xml$', item.filename):
                xml = data.decode('utf-8')
                # Extract axId and crossAx from existing catAx so axis links stay valid
                m = _catax_pat.search(xml)
                if m:
                    existing = m.group(0)
                    ax_id = re.search(r'<axId val="(\d+)"', existing)
                    cross_ax = re.search(r'<crossAx val="(\d+)"', existing)
                    ax_id_val = ax_id.group(1) if ax_id else "10"
                    cross_ax_val = cross_ax.group(1) if cross_ax else "100"

                    replacement = (
                        f'<catAx>'
                        f'<axId val="{ax_id_val}" />'
                        f'<scaling><orientation val="minMax" /></scaling>'
                        f'<delete val="0" />'
                        f'<axPos val="b" />'
                        f'<numFmt formatCode="General" sourceLinked="1" />'
                        f'<majorTickMark val="none" />'
                        f'<minorTickMark val="none" />'
                        f'<tickLblPos val="low" />'
                        f'<txPr>'
                        f'<a:bodyPr rot="-60000000" spcFirstLastPara="1" '
                        f'vertOverflow="ellipsis" vert="horz" wrap="square" '
                        f'anchor="ctr" anchorCtr="1"/>'
                        f'<a:lstStyle/>'
                        f'<a:p><a:pPr><a:defRPr sz="900" b="0" i="0" u="none" '
                        f'strike="noStrike" kern="1200" baseline="0"/>'
                        f'</a:pPr><a:endParaRPr lang="en-US"/></a:p>'
                        f'</txPr>'
                        f'<crossAx val="{cross_ax_val}" />'
                        f'<crosses val="min" />'
                        f'<auto val="1" />'
                        f'<lblAlgn val="ctr" />'
                        f'<lblOffset val="100" />'
                        f'<noMultiLvlLbl val="0" />'
                        f'</catAx>'
                    )
                    xml = _catax_pat.sub(replacement, xml, count=1)
                data = xml.encode('utf-8')
            zout.writestr(item, data)

    Path(xlsx_path).write_bytes(out_buf.getvalue())
    logger.info("Patched chart axes in %s", xlsx_path)


def _add_line_chart(ws, fiscal_years: list, latest_quarter: str = "Q4"):
    """Add LineChart matching Josh's FEMR Export Template 041526 exactly."""
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.colors import ColorChoice, SchemeColor
    from openpyxl.chart.shapes import GraphicalProperties

    chart = LineChart()
    chart.height = 15

    # Number of quarter columns — computed first so width can be set dynamically
    last_q_idx = QUARTERS.index(latest_quarter)
    num_quarters = (len(fiscal_years) - 1) * 4 + (last_q_idx + 1)

    # Dynamic width: ~1.5cm per quarter so all "Q1 FY20" labels fit on one row
    chart.width = max(25, num_quarters * 1.5)

    # No chart title — matches template exactly
    chart.legend.position = "r"
    chart.legend.overlay = False  # place legend outside plot area, not floating over it

    # Axis labels must be visible (delete=False) — without this openpyxl hides tick labels
    chart.y_axis.delete = False
    chart.y_axis.numFmt = '$#,##0.00'
    chart.x_axis.delete = False
    # Prevent Excel from splitting "Q1 FY20" into two-row multi-level categories
    chart.x_axis.noMultiLvlLbl = True
    first_data_col = 2   # col B
    last_data_col = first_data_col + num_quarters - 1

    # X-axis categories: row 51, cols B to last
    cats = Reference(ws, min_col=first_data_col, min_row=ROW_CUM_Q_LABELS,
                     max_col=last_data_col, max_row=ROW_CUM_Q_LABELS)

    # Series: each cumulative row becomes a line series
    series_defs = [
        ("Funds Committed",        ROW_CUM_COMMITTED),
        ("Obligated Funds",        ROW_CUM_OBLIGATED),
        ("Pre-Bill Expenditures",  ROW_CUM_EXPENDED),
        ("Budgeted Spend",         ROW_CUM_BUDGETED),
        ("Actual Expenditures",    ROW_CUM_ACTUAL),
    ]

    accent_colors = ["accent1", "accent2", "accent3", "accent4", "accent5"]

    for i, (title, row) in enumerate(series_defs):
        data_ref = Reference(ws, min_col=first_data_col, min_row=row,
                              max_col=last_data_col, max_row=row)
        chart.add_data(data_ref, from_rows=True, titles_from_data=False)
        s = chart.series[-1]
        s.tx = openpyxl_title(title)

        # Marker: circle, size 5, filled with theme accent color
        accent = accent_colors[i]
        scheme_clr = SchemeColor(val=accent)
        marker_fill = GraphicalProperties(solidFill=ColorChoice(schemeClr=scheme_clr))
        s.marker = Marker(symbol="circle", size=5, spPr=marker_fill)

        # Line: theme accent color, width 9525 EMU (~0.75pt)
        line_clr = ColorChoice(schemeClr=SchemeColor(val=accent))
        line = LineProperties(solidFill=line_clr, w=9525)
        s.graphicalProperties = GraphicalProperties(ln=line)

    chart.set_categories(cats)
    # Anchor at col B, row 59
    ws.add_chart(chart, f"B{ROW_CHART_START}")


def openpyxl_title(text: str):
    """Create a SeriesLabel with just a literal string value."""
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef
    sl = SeriesLabel(v=text)
    return sl


def build_tab(wb: Workbook, meta: dict, data: dict,
              fiscal_years: list, tab_name: str, latest_quarter: str = "Q4",
              pre_cumulative: Optional[dict] = None):
    """Build one tab per (sequence, rollup). meta includes all header fields."""
    # Unique sheet name
    sheet_name = tab_name
    counter = 2
    while sheet_name in wb.sheetnames:
        sheet_name = f"{tab_name}({counter})"
        counter += 1
    ws = wb.create_sheet(title=sheet_name[:31])

    _write_header_block(ws, meta)
    col_map = _write_col_headers(ws, fiscal_years, latest_quarter)

    last_year = fiscal_years[-1]
    last_q_idx = QUARTERS.index(latest_quarter)
    fye_q_pairs = [
        (_fye(y), q)
        for y in fiscal_years
        for q in (QUARTERS[:last_q_idx + 1] if y == last_year else QUARTERS)
    ]

    # ACTUALS section
    ac = ws.cell(row=ROW_ACTUALS_START, column=1, value="ACTUALS")
    ac.font = BOLD
    ac.fill = GREY_FILL
    for i, (label, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = ROW_ACTUALS_START + i
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "ACTUALS", api_account)

    ws.cell(row=ROW_ACTUALS_TOTAL, column=1, value="ACTUALS Total").fill = GREY_FILL
    ws.cell(row=ROW_ACTUALS_TOTAL, column=1).font = BOLD
    _write_section_total_row(ws, ROW_ACTUALS_TOTAL, ROW_ACTUALS_START, col_map)

    # BUDGETED section
    bc = ws.cell(row=ROW_BUDGET_START, column=1, value="BUDGETED")
    bc.font = BOLD
    bc.fill = GREY_FILL
    for i, (label, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = ROW_BUDGET_START + i
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "BUDGETED", api_account)

    ws.cell(row=ROW_BUDGET_TOTAL, column=1, value="BUDGETED Total").fill = GREY_FILL
    ws.cell(row=ROW_BUDGET_TOTAL, column=1).font = BOLD
    _write_section_total_row(ws, ROW_BUDGET_TOTAL, ROW_BUDGET_START, col_map)

    # CONTRACTING section
    cc = ws.cell(row=ROW_COMMITTED, column=1, value="CONTRACTING")
    cc.font = BOLD
    cc.fill = GREY_FILL
    ws.cell(row=ROW_COMMITTED, column=2, value="Committed")
    ws.cell(row=ROW_OBLIGATED, column=2, value="Obligated")
    ws.cell(row=ROW_EXPENDED,  column=2, value="Expended")
    ws.cell(row=ROW_REMAINING_CASH, column=2, value="Remaining Cash").font = BOLD

    _write_data_row(ws, ROW_COMMITTED, data, fye_q_pairs, col_map, "CONTRACTING", "Committed")
    _write_data_row(ws, ROW_OBLIGATED, data, fye_q_pairs, col_map, "CONTRACTING", "Obligated")
    _write_data_row(ws, ROW_EXPENDED,  data, fye_q_pairs, col_map, "CONTRACTING", "Expended")
    _write_remaining_cash_row(ws, col_map)

    # Cumulative section
    _write_cumulative_section(ws, fiscal_years, col_map, latest_quarter, pre_cumulative)

    # Chart
    _add_line_chart(ws, fiscal_years, latest_quarter)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 32
    for col_idx in col_map.values():
        if isinstance(col_idx, int):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "C17"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def tab_name_for(sequence: str, meta: dict, multi: bool) -> str:
    """
    Tab name rules:
      - Single identifier: use sequence
      - Multi-identifier: use 'SEQ R<rollup_display>'
    For orphans, rollup_display falls back to project_number.
    """
    if not multi:
        return sequence[:31]
    rollup_display = meta.get("rollup") or "NA"
    return f"{sequence} R{rollup_display}"[:31]


def process_sequence(wb: Workbook, seq_entry: dict, fiscal_years: list,
                     workers: int, preloaded_metadata: Optional[dict] = None,
                     latest_quarter: str = "Q4") -> int:
    """
    Process one sequence from the mapping file.
    May produce multiple tabs if the sequence has multiple identifiers
    (rollups + orphans).

    Returns number of tabs created.
    """
    sequence = seq_entry["sequence"]

    # Get all identifiers for this sequence from preloaded data (or fetch fallback)
    if preloaded_metadata is not None and sequence in preloaded_metadata:
        metas = preloaded_metadata[sequence]
    else:
        metas = fetch_sequence_identifiers(sequence)

    if not metas:
        logger.warning("No metadata for %s — skipping", sequence)
        return 0

    multi = len(metas) > 1

    if not multi:
        # Fast path: single identifier → use netamount API
        meta = metas[0]
        name = tab_name_for(sequence, meta, multi=False)
        logger.info("  [single-id] %s → tab %s", sequence, name)
        data = fetch_financials_single(sequence, fiscal_years, workers=workers,
                                       latest_quarter=latest_quarter)
        pre_cumulative = fetch_pre_fy2020_single(sequence, workers=workers)
        build_tab(wb, meta, data, fiscal_years, name, latest_quarter, pre_cumulative)
        return 1

    # Multi-identifier path: use MV to differentiate rollups vs orphans
    logger.info("  [multi-id] %s has %d identifiers: %s",
                sequence,
                len(metas),
                [(m.get("class"), m.get("rollup")) for m in metas])
    per_id_data = fetch_financials_multi(sequence, fiscal_years, workers=workers,
                                          latest_quarter=latest_quarter)
    pre_per_id = fetch_pre_fy2020_multi(sequence, workers=workers)

    tabs_written = 0
    for meta in metas:
        id_tuple = ("orphan", meta["project_number"]) if meta["is_orphan"] \
                   else ("rollup", meta["rollup_num"])
        data = per_id_data.get(id_tuple, {})
        if not data:
            logger.warning("  No data found for %s identifier %s — tab will be empty",
                           sequence, id_tuple)
        pre_cumulative = pre_per_id.get(id_tuple, {})
        name = tab_name_for(sequence, meta, multi=True)
        build_tab(wb, meta, data, fiscal_years, name, latest_quarter, pre_cumulative)
        tabs_written += 1

    return tabs_written


def run(mapping_path: str, output_prefix: str,
        single_sequence: Optional[str] = None,
        group_filter: Optional[str] = None,
        workers: int = 20,
        fiscal_year_end: int = FISCAL_YEAR_END_DEFAULT,
        skip_preload: bool = False,
        latest_quarter: Optional[str] = None,
        split_size: int = SPLIT_SIZE):

    logger.info("=== FEMR NetSuite Report Generator (v15) ===")

    # Detect latest quarter with data (unless overridden via --latest-quarter)
    if latest_quarter is None:
        fiscal_year_end, latest_quarter = detect_latest_fiscal_quarter()
    logger.info("Quarter range: Q1 FY%d → %s FY%d",
                FISCAL_YEAR_START, latest_quarter, fiscal_year_end)

    fiscal_years = list(range(FISCAL_YEAR_START, fiscal_year_end + 1))
    logger.info("Fiscal years: %s", fiscal_years)

    registry = load_sequence_registry(mapping_path)
    logger.info("Loaded %d unique sequences from mapping file", len(registry))

    if single_sequence:
        registry = [s for s in registry if s["sequence"] == single_sequence]
        if not registry:
            logger.error("Sequence %s not found in mapping", single_sequence)
            sys.exit(1)
        # Single-sequence mode: skip preload (just one metadata call)
        wb = Workbook()
        wb.remove(wb.active)
        for s in registry:
            process_sequence(wb, s, fiscal_years, workers,
                             preloaded_metadata=None, latest_quarter=latest_quarter)
        wb.save(output_prefix)
        _patch_chart_axes(output_prefix)
        logger.info("Saved %s", output_prefix)
        return

    if group_filter:
        registry = [s for s in registry if s["reporting_group"] == group_filter]
        logger.info("Filtered to %s: %d sequences", group_filter, len(registry))

    # Preload metadata via parallel per-sequence MV queries
    preloaded = None
    if not skip_preload:
        sequence_names = [s["sequence"] for s in registry]
        preloaded = preload_all_identifiers(sequence_names, workers=min(workers, 30))

    # Group registry by reporting_group
    by_group = {}
    for s in registry:
        by_group.setdefault(s["reporting_group"], []).append(s)

    total_tabs = 0
    total_files = 0
    for group, seqs in sorted(by_group.items()):
        # Split into chunks of SPLIT_SIZE
        chunks = [seqs[i:i + split_size] for i in range(0, len(seqs), split_size)]
        multi_file = len(chunks) > 1

        for chunk_idx, chunk in enumerate(chunks):
            if multi_file:
                start_num = chunk_idx * split_size + 1
                end_num   = start_num + len(chunk) - 1
                fname = f"{output_prefix}_{group.lower()}_{start_num:03d}-{end_num:03d}.xlsx"
            else:
                fname = f"{output_prefix}_{group.lower()}.xlsx"

            logger.info("\n--- Group: %s  chunk: %d/%d  sequences: %d  → %s ---",
                        group, chunk_idx + 1, len(chunks), len(chunk), fname)

            wb = Workbook()
            wb.remove(wb.active)

            for i, s in enumerate(chunk, 1):
                global_i = chunk_idx * split_size + i
                logger.info("[%d/%d] %s (%s)", global_i, len(seqs), s["sequence"], group)
                n = process_sequence(wb, s, fiscal_years, workers, preloaded,
                                     latest_quarter=latest_quarter)
                total_tabs += n
                if i % 10 == 0:
                    wb.save(fname)
                    logger.info("  Checkpoint saved (%d/%d in chunk)", i, len(chunk))

            wb.save(fname)
            _patch_chart_axes(fname)
            total_files += 1
            logger.info("Saved %s (total tabs so far: %d)", fname, total_tabs)

    logger.info("\nDone! %d total tabs across %d files.", total_tabs, total_files)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mapping", default="docs/Re_ FEMR /GROUP MAPPING.xlsx",
                        help="Path to GROUP MAPPING.xlsx")
    parser.add_argument("--output", "-o", default="femr_v10_report",
                        help="Output filename prefix (group name suffix added)")
    parser.add_argument("--sequence", "-s", default=None,
                        help="Single sequence only, writes to --output directly")
    parser.add_argument("--group", "-g", default=None,
                        help="Only generate for this reporting group (ADP, Comml, Internal, OGA, WFD)")
    parser.add_argument("--workers", "-w", type=int, default=20)
    parser.add_argument("--fy-end", type=int, default=FISCAL_YEAR_END_DEFAULT,
                        help="Last fiscal year to include (default 2026)")
    parser.add_argument("--skip-preload", action="store_true",
                        help="Skip bulk metadata preload (uses per-sequence fallback)")
    parser.add_argument("--latest-quarter", default=None,
                        help="Override auto-detected latest quarter e.g. 'Q2' (also set --fy-end)")
    parser.add_argument("--split-size", type=int, default=SPLIT_SIZE,
                        help=f"Max tabs per output file (default {SPLIT_SIZE}). Groups under this stay as one file.")
    args = parser.parse_args()

    if args.sequence and not args.output.endswith(".xlsx"):
        args.output = args.output + ".xlsx"

    run(mapping_path=args.mapping,
        output_prefix=args.output,
        single_sequence=args.sequence,
        group_filter=args.group,
        workers=args.workers,
        fiscal_year_end=args.fy_end,
        skip_preload=args.skip_preload,
        latest_quarter=args.latest_quarter,
        split_size=args.split_size)
