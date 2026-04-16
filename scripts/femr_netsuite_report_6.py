"""
FEMR NetSuite Report Generator (v6)
====================================
Matches the new FEMR Export Template 041526.xlsx layout.

Changes from v5:
  - Bulk metadata preload: 2 MV queries at startup (~90s) replace per-sequence
    metadata calls (~13 min → ~90s total for full run)
  - Multi-rollup sequences automatically detected from bulk data
  - Per-sequence fallback kept for edge cases not covered by bulk filter

Changes from v4:
  - NEW 11-row header block (REPORTING GROUP, SEQUENCE, ROLLUP #, PROJECT NAME,
    TECHNICAL POINT OF CONTACT, PROJECT TYPE, SERVICE, COLOR OF MONEY, ACRN,
    AVAILABLE FUNDS)
  - Data shifted to rows 17+ (was row 11)
  - Cumulative section moved to rows 51-56 (was 48-54)
  - Removed Revised Plan row
  - Removed govt awards / cash collected rows
  - Added Remaining Cash inside Contracting section (Obligated − Expended)
  - Rollup now used as API parameter — one tab per (sequence, rollup) pair
  - Dynamic mapping loaded from docs/Re_ FEMR /GROUP MAPPING.xlsx
  - Dynamic quarter range: FY2020 to latest quarter with data
  - Line chart embedded below cumulative section
  - All metadata fetched from /mv_femr_report/ ORDS filter query

Script supports all reporting groups: ADP, Comml, Internal, OGA, WFD.

Layout per tab:
  Row 1     : NextFlex
  Row 2     : REPORTING GROUP - TYPE A | value (ADP, Comml, ...)
  Row 3     : SEQUENCE | sequence code
  Row 4     : ROLLUP # | rollup number
  Row 5     : PROJECT NAME | legal name
  Row 6     : TECHNICAL POINT OF CONTACT | name
  Row 7     : PROJECT TYPE | type_b (comma-separated possible)
  Row 8     : SERVICE | service
  Row 9     : COLOR OF MONEY | (blank)
  Row 10    : ACRN | acrn
  Row 11    : AVAILABLE FUNDS | amount
  Row 12-14 : (blank)
  Row 15    : FYE headers (merged across 4 Q cols each)
  Row 16    : Q1 | Q2 | Q3 | Q4 ... | (Grand Total)
  Row 17-29 : Actuals (Labor Hours + 12 accounts)
  Row 30    : ACTUALS Total
  Row 31-43 : Budgeted (Labor Hours + 12 accounts)
  Row 44    : BUDGETED Total
  Row 45-48 : Contracting (Committed, Obligated, Expended, Remaining Cash)
  Row 49-50 : (blank)
  Row 51    : Q labels (Q1 FY20 ...)
  Row 52    : Total Committed (cumulative)
  Row 53    : Total Obligated (cumulative)
  Row 54    : Total Expended (cumulative)
  Row 55    : Budgeted Plan (cumulative)
  Row 56    : Actual (cumulative)
  Row 59+   : LineChart (5 series)

Usage:
    python femr_netsuite_report_6.py --sequence 2ADP001        # single tab test
    python femr_netsuite_report_6.py --group ADP               # all ADP sequences
    python femr_netsuite_report_6.py                           # all 5 groups
    python femr_netsuite_report_6.py --workers 30              # tune concurrency
"""

import argparse
import logging
import time
import urllib.request
import urllib.parse
import json
import sys
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── API ───────────────────────────────────────────────────────────────────────
BASE_URL = "https://g22673cc0c08b7a-oax2132513753.adb.us-ashburn-1.oraclecloudapps.com/ords/oax_user"
NETAMOUNT_URL = f"{BASE_URL}/femr/netamount/"
MV_URL = f"{BASE_URL}/mv_femr_report/"

# ── Fiscal scope ──────────────────────────────────────────────────────────────
FISCAL_YEAR_START = 2020
FISCAL_YEAR_END_DEFAULT = 2026   # used if MV lookup for latest year fails
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

# ── Account definitions ───────────────────────────────────────────────────────
# (row_label, api_account_name)
# api_account_name=None → statistical account, always blank
ACTUALS_BUDGET_ACCOUNTS = [
    ("Labor Hours statistical account", None),
    ("Labor Cost 5001",                 "5001 DIR : Direct Labor"),
    ("Fringe 5990",                     "5990 ALLO : Allo Fringe"),
    ("Travel 5004",                     "5004 DIR : Direct Travel"),
    ("Subcontracting 5005",             "5005 DIR : Subrecipient Costs"),
    ("Consulting 5002",                 "5002 DIR : Direct Consulting"),
    ("Equipment 5010",                  "5010 DIR : EQ & Materials (NO OH)"),
    ("Equipment 5008",                  "5008 DIR : Direct Equipment"),
    ("Other Direct Costs 5009",         "5009 DIR : Direct Other Costs"),
    ("Material 5003",                   "5003 DIR : Direct Materials"),
    ("Sub K Overhead 5992",             "5992 ALLO : Allo SubK OH"),
    ("Sub K Overhead 5993",             "5993 ALLO : DNU ALLO G and A OH WFD"),
    ("G&A 5991",                        "5991 ALLO : Allo G and A"),
]

CONTRACTING_ACCOUNTS = [
    ("Committed",  "Committed"),
    ("Obligated",  "Obligated"),
    ("Expended",   "Expended"),
]

# ── Layout constants ─────────────────────────────────────────────────────────
# Header block
ROW_FYE      = 15
ROW_Q        = 16
# Actuals
ROW_ACTUALS_START = 17   # Labor Hours row
ROW_ACTUALS_END   = 29   # G&A row
ROW_ACTUALS_TOTAL = 30
# Budgeted
ROW_BUDGET_START = 31
ROW_BUDGET_END   = 43
ROW_BUDGET_TOTAL = 44
# Contracting
ROW_COMMITTED      = 45
ROW_OBLIGATED      = 46
ROW_EXPENDED       = 47
ROW_REMAINING_CASH = 48
# Cumulative
ROW_CUM_Q_LABELS   = 51
ROW_CUM_COMMITTED  = 52
ROW_CUM_OBLIGATED  = 53
ROW_CUM_EXPENDED   = 54
ROW_CUM_BUDGETED   = 55
ROW_CUM_ACTUAL     = 56
# Chart
ROW_CHART_START    = 59

COL_DATA_START = 3  # col C

# ── Styles ────────────────────────────────────────────────────────────────────
FYE_GREEN  = PatternFill("solid", fgColor="C1FFB0")
FYE_BLUE   = PatternFill("solid", fgColor="A3D1FF")
GREY_FILL  = PatternFill("solid", fgColor="F5F4F2")

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
VCENTER_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

NUM_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# ─────────────────────────────────────────────────────────────────────────────
# MAPPING FILE LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_sequence_registry(mapping_path: str) -> list:
    """
    Load the GROUP MAPPING file and return a list of unique sequences with
    their reporting group. Returns: list of dicts with keys:
        sequence, reporting_group, type_a_category
    """
    wb = load_workbook(mapping_path, data_only=True, read_only=True)
    ws = wb["SEQUENCE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # Dedup on sequence — take the first occurrence
    seen = {}
    for row in rows:
        if not row or row[3] is None:
            continue
        rg, category, type_b, seq = row[0], row[1], row[2], row[3]
        if seq not in seen:
            seen[seq] = {
                "sequence": seq,
                "reporting_group": rg,
                "type_a_category": category,
            }
    return list(seen.values())


# ─────────────────────────────────────────────────────────────────────────────
# API FETCHING
# ─────────────────────────────────────────────────────────────────────────────

def _http_get(url: str, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=60) as r:
                return json.loads(r.read())
        except Exception as exc:
            if attempt == retries - 1:
                raise
            time.sleep(1.5 ** attempt)
    return {}


def _row_to_meta(i: dict) -> dict:
    """Extract metadata fields from an MV row."""
    return {
        "sequence":               i.get("display_sequence"),
        "rollup":                 i.get("display_rollup_num"),
        "project_legal_name":     i.get("display_project_legalname"),
        "technical_point_of_contact": i.get("technical_point_of_contact"),
        "acrn":                   i.get("display_acrn"),
        "available_funds":        i.get("available_funds") or 0,
        "type_b":                 i.get("display_type_b") or "",
        "service":                i.get("service") or i.get("display_type_c") or "",
        "reporting_group":        i.get("reporting_group_type_a"),
    }


def _bulk_metadata_fetch(filter_q: dict) -> dict:
    """Paginate MV with a tight filter. Returns {(seq, rollup): metadata}."""
    pairs = {}
    offset = 0
    page = 0
    while True:
        url = (f"{MV_URL}?q={urllib.parse.quote(json.dumps(filter_q))}"
               f"&limit=2000&offset={offset}")
        try:
            data = _http_get(url)
        except Exception as exc:
            logger.warning("Bulk metadata page %d failed: %s", page, exc)
            break
        items = data.get("items", [])
        page += 1
        for i in items:
            key = (i.get("display_sequence"), i.get("display_rollup_num"))
            if key not in pairs:
                pairs[key] = _row_to_meta(i)
        if not data.get("hasMore"):
            break
        offset += len(items)
        if page > 10:
            logger.warning("Bulk metadata hit 10-page safety cap")
            break
    return pairs


def preload_all_metadata() -> dict:
    """
    Preload metadata for all (sequence, rollup) pairs via 2 bulk MV queries.

    Strategy: combining ACTUALS + BUDGETED Q1 FY2020 for Labor Cost 5001
    covers essentially all active sequences. The ~90s one-time cost replaces
    ~2s per sequence (~13 min for 398 sequences).

    Returns: {(sequence, rollup): metadata_dict}
    """
    logger.info("Preloading metadata via bulk MV queries...")
    t0 = time.time()

    filters = [
        {"segment": "ACTUALS",  "account_name": "5001 DIR : Direct Labor",
         "fiscal_year_end": "FYE 9/30/2020", "fiscal_quarter": "Q1"},
        {"segment": "BUDGETED", "account_name": "5001 DIR : Direct Labor",
         "fiscal_year_end": "FYE 9/30/2020", "fiscal_quarter": "Q1"},
    ]

    all_pairs = {}
    for idx, f in enumerate(filters, 1):
        pairs = _bulk_metadata_fetch(f)
        new_count = 0
        for k, v in pairs.items():
            if k not in all_pairs:
                all_pairs[k] = v
                new_count += 1
        logger.info("  Filter %d: %d rows, %d new pairs (total: %d)",
                    idx, len(pairs), new_count, len(all_pairs))

    logger.info("Preloaded %d (sequence, rollup) pairs in %.1fs",
                len(all_pairs), time.time() - t0)
    return all_pairs


def fetch_sequence_metadata(sequence: str, preloaded: Optional[dict] = None) -> list:
    """
    Return list of metadata dicts (one per rollup) for a sequence.
    Uses preloaded bulk data if available. Falls back to per-sequence MV
    query for sequences not found in the bulk preload.
    """
    if preloaded is not None:
        matches = [v for (s, r), v in preloaded.items() if s == sequence]
        if matches:
            return matches
        # Fall through to per-sequence fetch if not in preloaded

    q = json.dumps({"display_sequence": sequence})
    url = f"{MV_URL}?q={urllib.parse.quote(q)}&limit=1"

    try:
        data = _http_get(url)
    except Exception as exc:
        logger.warning("Metadata fallback fetch failed for %s: %s", sequence, exc)
        return []

    items = data.get("items", [])
    if not items:
        return []

    return [_row_to_meta(items[0])]


def _fetch_netamount(sequence: str, rollup: Optional[str], fye: str, quarter: str,
                     segment: str, account_name: str, retries: int = 3) -> float:
    params = {
        "display_sequence": sequence,
        "fiscal_year_end":  fye,
        "fiscal_quarter":   quarter,
        "segment":          segment,
        "account_name":     account_name,
    }
    if rollup is not None:
        params["display_rollup_num"] = rollup
    url = f"{NETAMOUNT_URL}?{urllib.parse.urlencode(params)}"
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                data = json.loads(r.read())
            val = data.get("items", [{}])[0].get("total_netamount")
            return float(val) if val is not None else 0.0
        except Exception as exc:
            if attempt == retries - 1:
                logger.warning("FAILED %s/%s %s %s %s: %s",
                               sequence, rollup, fye, quarter, account_name, exc)
                return 0.0
            time.sleep(1.5 ** attempt)
    return 0.0


def _build_tasks(fiscal_years: list) -> list:
    """Build (fye, quarter, segment, account) tuples."""
    tasks = []
    for year in fiscal_years:
        fye = f"FYE 9/30/{year}"
        for quarter in QUARTERS:
            for _, account in ACTUALS_BUDGET_ACCOUNTS:
                if account is not None:
                    tasks.append((fye, quarter, "ACTUALS",  account))
                    tasks.append((fye, quarter, "BUDGETED", account))
            for _, account in CONTRACTING_ACCOUNTS:
                tasks.append((fye, quarter, "CONTRACTING", account))
    return tasks


def fetch_financials(sequence: str, rollup: Optional[str],
                     fiscal_years: list, workers: int = 20) -> dict:
    """Fetch all financial values for one (sequence, rollup). Returns dict keyed by task tuple."""
    tasks = _build_tasks(fiscal_years)
    results = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_netamount, sequence, rollup, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            key = future_map[future]
            results[key] = future.result()
    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────

def _fye(year: int) -> str:
    return f"FYE 9/30/{year}"


def _q_label(year: int, q: str) -> str:
    return f"{q} FY{str(year)[2:]}"


def _write_header_block(ws, meta: dict):
    """Write rows 1-11 header block."""
    fields = [
        ("NextFlex",                      ""),
        ("REPORTING GROUP - TYPE A",      meta.get("reporting_group") or ""),
        ("SEQUENCE",                      meta.get("sequence") or ""),
        ("ROLLUP #",                      meta.get("rollup") or ""),
        ("PROJECT NAME",                  meta.get("project_legal_name") or ""),
        ("TECHNICAL POINT OF CONTACT",    meta.get("technical_point_of_contact") or ""),
        ("PROJECT TYPE",                  meta.get("type_b") or ""),
        ("SERVICE",                       meta.get("service") or ""),
        ("COLOR OF MONEY",                ""),
        ("ACRN",                          meta.get("acrn") or ""),
        ("AVAILABLE FUNDS",               meta.get("available_funds") or 0),
    ]
    for r, (label, value) in enumerate(fields, start=1):
        la = ws.cell(row=r, column=1, value=label)
        la.font = BOLD
        ws.cell(row=r, column=2, value=value)


def _write_col_headers(ws, fiscal_years: list) -> dict:
    """Write FYE (row 15) and Q (row 16) headers, return col_map."""
    col = COL_DATA_START
    col_map = {}
    # Alternating fills: FY2020 green, FY2021 blue, etc.
    fills = [FYE_GREEN, FYE_BLUE]

    for idx, year in enumerate(fiscal_years):
        fye = _fye(year)
        fill = fills[idx % 2]
        # Merge FYE across 4 Q cols
        ws.merge_cells(start_row=ROW_FYE, start_column=col,
                       end_row=ROW_FYE,   end_column=col + 3)
        c = ws.cell(row=ROW_FYE, column=col, value=fye)
        c.alignment = CENTER
        c.font = BOLD
        c.fill = fill
        # Fill all 4 Q cols with same color
        for qi in range(4):
            ws.cell(row=ROW_FYE, column=col + qi).fill = fill
            qc = ws.cell(row=ROW_Q, column=col + qi, value=QUARTERS[qi])
            qc.alignment = CENTER
            qc.font = BOLD
            qc.fill = fill
            col_map[(fye, QUARTERS[qi])] = col + qi
        col += 4

    # Grand Total column
    gt = ws.cell(row=ROW_FYE, column=col, value="Grand Total")
    gt.alignment = CENTER
    gt.font = BOLD
    gt.fill = GREY_FILL
    ws.cell(row=ROW_Q, column=col).fill = GREY_FILL
    col_map["total"] = col

    return col_map


def _write_data_row(ws, row: int, data: dict, fye_q_pairs: list,
                    col_map: dict, segment: str, api_account: Optional[str]):
    """Write a single data row with values + Grand Total SUM formula."""
    first_col = None
    last_col = None
    for (fye, q) in fye_q_pairs:
        col = col_map.get((fye, q))
        if col is None:
            continue
        if first_col is None:
            first_col = col
        last_col = col
        amount = 0.0 if api_account is None else data.get((fye, q, segment, api_account), 0.0)
        if amount != 0.0:
            c = ws.cell(row=row, column=col, value=amount)
            c.number_format = NUM_FMT
            c.alignment = RIGHT

    # Grand Total
    total_col = col_map.get("total")
    if total_col and first_col and last_col:
        fl = get_column_letter(first_col)
        ll = get_column_letter(last_col)
        tc = ws.cell(row=row, column=total_col,
                     value=f"=SUM({fl}{row}:{ll}{row})")
        tc.number_format = NUM_FMT
        tc.alignment = RIGHT
        tc.fill = GREY_FILL


def _write_section_total_row(ws, row: int, data_start_row: int, col_map: dict):
    """Write a SUM total row across account rows (skips Labor Hours row)."""
    sum_start = data_start_row + 1  # skip Labor Hours
    for col_idx in col_map.values():
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=row, column=col_idx,
                    value=f"=SUM({col_letter}{sum_start}:{col_letter}{row - 1})")
        c.number_format = NUM_FMT
        c.alignment = RIGHT
        c.fill = GREY_FILL


def _write_remaining_cash_row(ws, col_map: dict):
    """Remaining Cash = Obligated - Expended per quarter."""
    total_col = col_map.get("total")
    first_col = None
    last_col = None
    for key, col in col_map.items():
        if not isinstance(key, tuple):
            continue
        if first_col is None or col < first_col:
            first_col = col
        if last_col is None or col > last_col:
            last_col = col
        cl = get_column_letter(col)
        c = ws.cell(row=ROW_REMAINING_CASH, column=col,
                    value=f"={cl}{ROW_OBLIGATED}-{cl}{ROW_EXPENDED}")
        c.number_format = NUM_FMT
        c.alignment = RIGHT

    if total_col and first_col and last_col:
        fl = get_column_letter(first_col)
        ll = get_column_letter(last_col)
        tc = ws.cell(row=ROW_REMAINING_CASH, column=total_col,
                     value=f"=SUM({fl}{ROW_REMAINING_CASH}:{ll}{ROW_REMAINING_CASH})")
        tc.number_format = NUM_FMT
        tc.alignment = RIGHT
        tc.fill = GREY_FILL


def _write_cumulative_section(ws, fiscal_years: list, col_map: dict):
    """
    Rows 51-56: cumulative running totals.
    Row 51: Q labels starting at col B
    Row 52: Total Committed (cumulative from row 45)
    Row 53: Total Obligated (cumulative from row 46)
    Row 54: Total Expended (cumulative from row 47)
    Row 55: Budgeted Plan (cumulative from row 44)
    Row 56: Actual (cumulative from row 30)
    """
    # Row 51: Q labels - start at col B (col 2) per new template
    col_b = 2
    for year in fiscal_years:
        for q in QUARTERS:
            c = ws.cell(row=ROW_CUM_Q_LABELS, column=col_b, value=_q_label(year, q))
            c.alignment = CENTER
            c.font = BOLD
            col_b += 1

    formula_defs = [
        (ROW_CUM_COMMITTED, "Total Committed",  ROW_COMMITTED),
        (ROW_CUM_OBLIGATED, "Total Obligated",  ROW_OBLIGATED),
        (ROW_CUM_EXPENDED,  "Total Expended",   ROW_EXPENDED),
        (ROW_CUM_BUDGETED,  "Budgeted Plan",    ROW_BUDGET_TOTAL),
        (ROW_CUM_ACTUAL,    "Actual",           ROW_ACTUALS_TOTAL),
    ]

    for (frow, label, src_row) in formula_defs:
        ws.cell(row=frow, column=1, value=label).font = BOLD

        # Column B holds first quarter; each subsequent adds to prev cumulative
        prev_col_letter = None
        col_b = 2
        for year in fiscal_years:
            for q in QUARTERS:
                # The source row data is in col_map at (fye, q) col
                src_col = col_map.get((_fye(year), q))
                if src_col is None:
                    col_b += 1
                    continue
                src_col_letter = get_column_letter(src_col)
                this_col_letter = get_column_letter(col_b)

                if prev_col_letter is None:
                    formula = f"={src_col_letter}{src_row}"
                else:
                    formula = f"={src_col_letter}{src_row}+{prev_col_letter}{frow}"

                c = ws.cell(row=frow, column=col_b, value=formula)
                c.number_format = NUM_FMT
                c.alignment = RIGHT
                prev_col_letter = this_col_letter
                col_b += 1


def _add_line_chart(ws, fiscal_years: list):
    """Add LineChart below cumulative data, 5 series matching template."""
    chart = LineChart()
    chart.style = 2
    chart.height = 10
    chart.width = 22

    # Number of quarter columns = total cumulative columns
    num_quarters = len(fiscal_years) * 4
    first_data_col = 2   # col B
    last_data_col = first_data_col + num_quarters - 1

    # X-axis categories: row 51, cols B to last
    cats = Reference(ws, min_col=first_data_col, min_row=ROW_CUM_Q_LABELS,
                     max_col=last_data_col, max_row=ROW_CUM_Q_LABELS)

    # Series: each cumulative row becomes a line series
    series_defs = [
        ("Funds Committed",        ROW_CUM_COMMITTED),
        ("Obligated Funds",        ROW_CUM_OBLIGATED),
        ("Pre-Bill Expenditures",  ROW_CUM_EXPENDED),
        ("Budgeted Spend",         ROW_CUM_BUDGETED),
        ("Actual Expenditures",    ROW_CUM_ACTUAL),
    ]

    for (title, row) in series_defs:
        data_ref = Reference(ws, min_col=first_data_col, min_row=row,
                              max_col=last_data_col, max_row=row)
        chart.add_data(data_ref, from_rows=True, titles_from_data=False)
        chart.series[-1].tx = openpyxl_title(title)
        chart.series[-1].marker = Marker(symbol="circle", size=6)

    chart.set_categories(cats)
    # Anchor at col B, row 59
    ws.add_chart(chart, f"B{ROW_CHART_START}")


def openpyxl_title(text: str):
    """Create a SeriesLabel with just a literal string value."""
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.data_source import StrRef
    sl = SeriesLabel(v=text)
    return sl


def build_tab(wb: Workbook, meta: dict, data: dict,
              fiscal_years: list, tab_name: str):
    """Build one tab per (sequence, rollup). meta includes all header fields."""
    # Unique sheet name
    sheet_name = tab_name
    counter = 2
    while sheet_name in wb.sheetnames:
        sheet_name = f"{tab_name}({counter})"
        counter += 1
    ws = wb.create_sheet(title=sheet_name[:31])

    _write_header_block(ws, meta)
    col_map = _write_col_headers(ws, fiscal_years)

    fye_q_pairs = [(_fye(y), q) for y in fiscal_years for q in QUARTERS]

    # ACTUALS section
    ac = ws.cell(row=ROW_ACTUALS_START, column=1, value="ACTUALS")
    ac.font = BOLD
    ac.fill = GREY_FILL
    for i, (label, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = ROW_ACTUALS_START + i
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "ACTUALS", api_account)

    ws.cell(row=ROW_ACTUALS_TOTAL, column=1, value="ACTUALS Total").fill = GREY_FILL
    ws.cell(row=ROW_ACTUALS_TOTAL, column=1).font = BOLD
    _write_section_total_row(ws, ROW_ACTUALS_TOTAL, ROW_ACTUALS_START, col_map)

    # BUDGETED section
    bc = ws.cell(row=ROW_BUDGET_START, column=1, value="BUDGETED")
    bc.font = BOLD
    bc.fill = GREY_FILL
    for i, (label, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = ROW_BUDGET_START + i
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "BUDGETED", api_account)

    ws.cell(row=ROW_BUDGET_TOTAL, column=1, value="BUDGETED Total").fill = GREY_FILL
    ws.cell(row=ROW_BUDGET_TOTAL, column=1).font = BOLD
    _write_section_total_row(ws, ROW_BUDGET_TOTAL, ROW_BUDGET_START, col_map)

    # CONTRACTING section
    cc = ws.cell(row=ROW_COMMITTED, column=1, value="CONTRACTING")
    cc.font = BOLD
    cc.fill = GREY_FILL
    ws.cell(row=ROW_COMMITTED, column=2, value="Committed")
    ws.cell(row=ROW_OBLIGATED, column=2, value="Obligated")
    ws.cell(row=ROW_EXPENDED,  column=2, value="Expended")
    ws.cell(row=ROW_REMAINING_CASH, column=2, value="Remaining Cash").font = BOLD

    _write_data_row(ws, ROW_COMMITTED, data, fye_q_pairs, col_map, "CONTRACTING", "Committed")
    _write_data_row(ws, ROW_OBLIGATED, data, fye_q_pairs, col_map, "CONTRACTING", "Obligated")
    _write_data_row(ws, ROW_EXPENDED,  data, fye_q_pairs, col_map, "CONTRACTING", "Expended")
    _write_remaining_cash_row(ws, col_map)

    # Cumulative section
    _write_cumulative_section(ws, fiscal_years, col_map)

    # Chart
    _add_line_chart(ws, fiscal_years)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 32
    for col_idx in col_map.values():
        if isinstance(col_idx, int):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = "C17"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def tab_name_for(meta: dict) -> str:
    """Generate tab name: 'SEQ' or 'SEQ R<rollup>' if multiple rollups exist."""
    seq = meta["sequence"]
    rollup = meta.get("rollup")
    if rollup:
        return f"{seq} R{rollup}"[:31]
    return seq[:31]


def process_sequence(wb: Workbook, seq_entry: dict, fiscal_years: list,
                     workers: int, preloaded_metadata: Optional[dict] = None) -> int:
    """
    Fetch metadata + all rollups, generate one tab per (seq, rollup).
    Returns number of tabs created.
    """
    sequence = seq_entry["sequence"]
    rollups_meta = fetch_sequence_metadata(sequence, preloaded=preloaded_metadata)
    if not rollups_meta:
        logger.warning("No metadata for %s — skipping", sequence)
        return 0

    multi = len(rollups_meta) > 1
    for meta in rollups_meta:
        rollup = meta.get("rollup")
        # If only one rollup, use just the sequence as tab name; otherwise include rollup
        if multi:
            name = f"{sequence} R{rollup or 'NA'}"[:31]
        else:
            name = sequence[:31]

        logger.info("  Fetching financials for %s / rollup=%s", sequence, rollup)
        data = fetch_financials(sequence, rollup, fiscal_years, workers=workers)
        build_tab(wb, meta, data, fiscal_years, name)

    return len(rollups_meta)


def run(mapping_path: str, output_prefix: str,
        single_sequence: Optional[str] = None,
        group_filter: Optional[str] = None,
        workers: int = 20,
        fiscal_year_end: int = FISCAL_YEAR_END_DEFAULT,
        skip_preload: bool = False):

    logger.info("=== FEMR NetSuite Report Generator (v6) ===")

    fiscal_years = list(range(FISCAL_YEAR_START, fiscal_year_end + 1))
    logger.info("Fiscal years: %s", fiscal_years)

    registry = load_sequence_registry(mapping_path)
    logger.info("Loaded %d unique sequences from mapping file", len(registry))

    # Preload metadata via bulk MV queries
    preloaded = None if skip_preload else preload_all_metadata()

    if single_sequence:
        registry = [s for s in registry if s["sequence"] == single_sequence]
        if not registry:
            logger.error("Sequence %s not found in mapping", single_sequence)
            sys.exit(1)
        # Write to single file
        wb = Workbook()
        wb.remove(wb.active)
        for s in registry:
            process_sequence(wb, s, fiscal_years, workers, preloaded)
        wb.save(output_prefix)
        logger.info("Saved %s", output_prefix)
        return

    if group_filter:
        registry = [s for s in registry if s["reporting_group"] == group_filter]
        logger.info("Filtered to %s: %d sequences", group_filter, len(registry))

    # Group registry by reporting_group
    by_group = {}
    for s in registry:
        by_group.setdefault(s["reporting_group"], []).append(s)

    total_tabs = 0
    for group, seqs in sorted(by_group.items()):
        fname = f"{output_prefix}_{group.lower()}.xlsx"
        logger.info("\n--- Group: %s  sequences: %d  → %s ---",
                    group, len(seqs), fname)

        wb = Workbook()
        wb.remove(wb.active)

        for i, s in enumerate(seqs, 1):
            logger.info("[%d/%d] %s (%s)", i, len(seqs), s["sequence"], group)
            n = process_sequence(wb, s, fiscal_years, workers, preloaded)
            total_tabs += n
            if i % 10 == 0:
                wb.save(fname)
                logger.info("  Checkpoint saved (%d/%d)", i, len(seqs))

        wb.save(fname)
        logger.info("Saved %s (tabs so far: %d)", fname, total_tabs)

    logger.info("\nDone! %d total tabs across %d files.", total_tabs, len(by_group))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mapping", default="docs/Re_ FEMR /GROUP MAPPING.xlsx",
                        help="Path to GROUP MAPPING.xlsx")
    parser.add_argument("--output", "-o", default="femr_v6_report",
                        help="Output filename prefix (group name suffix added)")
    parser.add_argument("--sequence", "-s", default=None,
                        help="Single sequence only, writes to --output directly")
    parser.add_argument("--group", "-g", default=None,
                        help="Only generate for this reporting group (ADP, Comml, Internal, OGA, WFD)")
    parser.add_argument("--workers", "-w", type=int, default=20)
    parser.add_argument("--fy-end", type=int, default=FISCAL_YEAR_END_DEFAULT,
                        help="Last fiscal year to include (default 2026)")
    parser.add_argument("--skip-preload", action="store_true",
                        help="Skip bulk metadata preload (uses per-sequence fallback)")
    args = parser.parse_args()

    if args.sequence and not args.output.endswith(".xlsx"):
        args.output = args.output + ".xlsx"

    run(mapping_path=args.mapping,
        output_prefix=args.output,
        single_sequence=args.sequence,
        group_filter=args.group,
        workers=args.workers,
        fiscal_year_end=args.fy_end,
        skip_preload=args.skip_preload)
