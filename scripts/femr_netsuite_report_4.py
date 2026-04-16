"""
FEMR NetSuite Report Generator (v4)
====================================
Generates multi-tab Excel workbook matching FEMR_for_NetSuite.xlsx template exactly.

Changes from v3:
  - Fiscal year range corrected: FY2020–FY2026 (per Josh/NSAW, not FY2018–FY2025)
  - Formulas section (rows 49–54) now uses cumulative running totals
    (first col = direct ref, subsequent = current_quarter + prev_cumulative)
  - Header block: 8 rows matching template (added External Approver 1,
    labels "Project Type" / "Service" instead of "Type B" / "Type C")
  - Added per-row Total column with =SUM() across all quarters
  - All rows kept even if blank (per Shoma: "keep it blank, don't skip")

Row layout per tab (matches ADP 33 template):
  Row 1     : NextFlex
  Row 2     : ADP  | number | Rollup (if applicable)
  Row 3     : Project Name | value
  Row 4     : External Approver 1 | value (blank — not in API yet)
  Row 5     : Project Type | value (Type B)
  Row 6     : Service | value (Type C)
  Row 7     : Color of Money | value (blank — not in API)
  Row 8     : ACRNs | value
  Row 9     : (blank) | (blank) | FYE 9/30/2020 ... (merged per FYE x 4 cols)
  Row 10    : (blank) | (blank) | Q1 | Q2 | Q3 | Q4 | Q1 ...
  Row 11-24 : Actuals section (A11:A24 merged = "Actuals")
              B = account name, C onwards = data
  Row 25-38 : Budgeted section (A25:A38 merged = "Budgeted")
  Row 39-41 : Contracting section (A39:A41 merged = "Contracting")
              Committed / Obligated / Expended
  Row 42    : (blank)
  Row 43    : (blank) | govt awards
  Row 44    : (blank) | govt obligated
  Row 45    : (blank) | cash collected
  Row 46    : add calc field | remaining cash
  Row 47    : (blank)
  Row 48    : (blank) | (blank) | Q1 FY20 | Q2 FY20 ...
  Row 49    : formulas | Total Committed | cumulative values...
  Row 50    : (blank) | Total Obligated
  Row 51    : (blank) | Total Expended
  Row 52    : (blank) | Budgeted Plan
  Row 53    : (blank) | Revised Plan
  Row 54    : (blank) | Actual

Usage:
    pip install openpyxl
    python femr_netsuite_report_4.py --sequence 2ADP001        # single tab test
    python femr_netsuite_report_4.py --ca2-only                # CA2 ADPs only
    python femr_netsuite_report_4.py                           # all sequences
    python femr_netsuite_report_4.py --workers 30              # tune concurrency
"""

import argparse
import logging
import time
import urllib.request
import urllib.parse
import json
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── API ───────────────────────────────────────────────────────────────────────
NETAMOUNT_URL = (
    "https://g22673cc0c08b7a-oax2132513753.adb.us-ashburn-1.oraclecloudapps.com"
    "/ords/oax_user/femr/netamount/"
)

# ── Fiscal scope (corrected: FY2020–FY2026 per Josh/NSAW) ────────────────────
FISCAL_YEARS = [2020, 2021, 2022, 2023, 2024, 2025, 2026]
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

# ── Account definitions ───────────────────────────────────────────────────────
# (row_label, account_code, api_account_name)
# api_account_name=None → statistical/no API data, always 0
ACTUALS_BUDGET_ACCOUNTS = [
    ("Labor Hours statistical account", None,   None),
    ("Labor Cost 5001",                 "5001",  "5001 DIR : Direct Labor"),
    ("Fringe 5990",                     "5990",  "5990 ALLO : Allo Fringe"),
    ("Travel 5004",                     "5004",  "5004 DIR : Direct Travel"),
    ("Subcontracting 5005",             "5005",  "5005 DIR : Subrecipient Costs"),
    ("Consulting 5002",                 "5002",  "5002 DIR : Direct Consulting"),
    ("Equipment 5010",                  "5010",  "5010 DIR : EQ & Materials (NO OH)"),
    ("Equipment 5008",                  "5008",  "5008 DIR : Direct Equipment"),
    ("Other Direct Costs 5009",         "5009",  "5009 DIR : Direct Other Costs"),
    ("Material 5003",                   "5003",  "5003 DIR : Direct Materials"),
    ("Sub K Overhead 5992",             "5992",  "5992 ALLO : Allo SubK OH"),
    ("Sub K Overhead 5993",             "5993",  "5993 ALLO : DNU ALLO G and A OH WFD"),
    ("G&A 5991",                        "5991",  "5991 ALLO : Allo G and A"),
]

CONTRACTING_ACCOUNTS = [
    ("Committed",  "Committed"),
    ("Obligated",  "Obligated"),
    ("Expended",   "Expended"),
]

# ── ADP Registry (224 sequences) ──────────────────────────────────────────────
ADP_REGISTRY = {
    "1ADP001": {"adp": 1,   "name": "CA1 ADP1 ADP2 ADP6 ADP11",                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP002": {"adp": 2,   "name": "CA1 ADP2",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP003": {"adp": 3,   "name": "CA1 ADP3",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP004": {"adp": 4,   "name": "CA1 ADP4a",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP005": {"adp": 5,   "name": "CA1 ADP5",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP006": {"adp": 6,   "name": "CA1 ADP6",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP007": {"adp": 7,   "name": "CA1 ADP7",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP008": {"adp": 8,   "name": "CA1 ADP8",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP009": {"adp": 9,   "name": "CA1 ADP9",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP010": {"adp": 10,  "name": "CA1 ADP10",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP011": {"adp": 11,  "name": "CA1 ADP11",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP012": {"adp": 12,  "name": "CA1 ADP12",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP013": {"adp": 13,  "name": "CA1 ADP13",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP014": {"adp": 14,  "name": "CA1 ADP14",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP015": {"adp": 15,  "name": "CA1 ADP15",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP016": {"adp": 16,  "name": "CA1 ADP16",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP017": {"adp": 17,  "name": "CA1 ADP17",                                   "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "1ADP018": {"adp": 18,  "name": "CA1 ADP18",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP019": {"adp": 19,  "name": "CA1 ADP19",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP020": {"adp": 20,  "name": "CA1 ADP20",                                   "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "1ADP021": {"adp": 21,  "name": "CA1 ADP21",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP022": {"adp": 22,  "name": "CA1 ADP22",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP023": {"adp": 23,  "name": "CA1 ADP23",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP024": {"adp": 24,  "name": "CA1 ADP24",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP025": {"adp": 25,  "name": "CA1 ADP25",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP026": {"adp": 26,  "name": "CA1 ADP26",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP027": {"adp": 27,  "name": "CA1 ADP27",                                   "type_b": "ADP-EWD",                                 "type_c": "Other DOD",       "rollup": None},
    "1ADP028": {"adp": 28,  "name": "CA1 ADP28",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP029": {"adp": 29,  "name": "CA1 ADP29",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP030": {"adp": 30,  "name": "CA1 ADP30",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP031": {"adp": 31,  "name": "CA1 ADP31",                                   "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "1ADP032": {"adp": 32,  "name": "CA1 ADP32",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP034": {"adp": 34,  "name": "CA1 ADP34",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP035": {"adp": 35,  "name": "CA1 ADP35",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP036": {"adp": 36,  "name": "CA1 ADP36",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP037": {"adp": 37,  "name": "CA1 ADP37",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP038": {"adp": 38,  "name": "CA1 ADP38",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP039": {"adp": 39,  "name": "CA1 ADP39",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP040": {"adp": 40,  "name": "CA1 ADP40",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP041": {"adp": 41,  "name": "CA1 ADP41",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP042": {"adp": 42,  "name": "CA1 ADP42",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP043": {"adp": 43,  "name": "CA1 ADP43",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP044": {"adp": 44,  "name": "CA1 ADP44",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP045": {"adp": 45,  "name": "CA1 ADP45",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP046": {"adp": 46,  "name": "CA1 ADP47",                                   "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "1ADP047": {"adp": 47,  "name": "CA1 ADP48",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP048": {"adp": 48,  "name": "CA1 ADP46",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP049": {"adp": 49,  "name": "CA1 ADP49",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP050": {"adp": 50,  "name": "CA1 ADP50",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP051": {"adp": 51,  "name": "CA1 ADP51",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP052": {"adp": 52,  "name": "CA1 ADP52",                                   "type_b": "ADP-Tech",                                "type_c": "Other Gov",       "rollup": None},
    "1ADP053": {"adp": 53,  "name": "CA1 ADP53",                                   "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "1ADP054": {"adp": 54,  "name": "CA1 ADP54",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP055": {"adp": 55,  "name": "CA1 ADP55",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP056": {"adp": 56,  "name": "CA1 ADP56",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP057": {"adp": 57,  "name": "CA1 ADP57",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP058": {"adp": 58,  "name": "CA1 ADP58",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP059": {"adp": 59,  "name": "CA1 ADP59",                                   "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "1ADP060": {"adp": 60,  "name": "CA1 ADP60",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1ADP061": {"adp": 61,  "name": "CA1 ADP61",                                   "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "1PC001":  {"adp": "PC1",  "name": "PC 0 IQM",                                 "type_b": "PC",  "type_c": "Other Gov",       "rollup": None},
    "1PC002":  {"adp": "PC2",  "name": "PC 1 Cal Poly",                             "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC003":  {"adp": "PC3",  "name": "PC 1 Binghamton Univ",                      "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC004":  {"adp": "PC4",  "name": "PC 1 Univ of MA Amherst",                   "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC005":  {"adp": "PC5",  "name": "PC 1 Univ of MA Lowell",                    "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC006":  {"adp": "PC6",  "name": "PC 1 UTRC",                                 "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC007":  {"adp": "PC7",  "name": "PC 1 PARC",                                 "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC008":  {"adp": "PC8",  "name": "PC 1 Purdue",                               "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC009":  {"adp": "PC9",  "name": "PC 1 UC Berkeley",                          "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC010":  {"adp": "PC10", "name": "PC 2 Uniqarta Inc",                         "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC011":  {"adp": "PC11", "name": "PC 2 Universal Instruments",                "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC012":  {"adp": "PC12", "name": "PC 2 Meyer Burger US",                      "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC013":  {"adp": "PC13", "name": "PC 2 Sensor Films Inc",                     "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC014":  {"adp": "PC14", "name": "PC 2 Boeing 2.3",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC015":  {"adp": "PC15", "name": "PC 2 Lockheed Martin",                      "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC016":  {"adp": "PC16", "name": "PC 2 Hewlett Packard Labs",                 "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC017":  {"adp": "PC17", "name": "PC 2 Georgia Tech",                         "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC018":  {"adp": "PC18", "name": "PC 2 U of Mass Lowell 2.0",                 "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC019":  {"adp": "PC19", "name": "PC 2 Auburn University",                    "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC020":  {"adp": "PC20", "name": "PC 2 SI2 Technologies",                     "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC021":  {"adp": "PC21", "name": "PC 2 Boeing 2.6",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC022":  {"adp": "PC22", "name": "PC 2 American Semiconductor",               "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC023":  {"adp": "PC23", "name": "PC 2 Purdue Univ 2.0",                      "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC024":  {"adp": "PC24", "name": "PC 2 Binghamton Univ Proj 4",               "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC025":  {"adp": "PC25", "name": "PC 2 Lorain College",                       "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC026":  {"adp": "PC26", "name": "PC 3 Lockheed Martin 3.1",                  "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC027":  {"adp": "PC27", "name": "PC 3 Boeing PPFHE",                         "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC028":  {"adp": "PC28", "name": "PC 3 MicroConnex",                          "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC029":  {"adp": "PC29", "name": "PC 3 Boeing CMSA",                          "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC030":  {"adp": "PC30", "name": "PC 3 GA Tech Lockheed Martin 3.6",          "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC031":  {"adp": "PC31", "name": "PC 3 Binghamton Univ 3.0",                  "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC032":  {"adp": "PC32", "name": "Binghamton Univ. 3.0",                      "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC033":  {"adp": "PC33", "name": "PC 3 Epicore 3.0",                          "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC034":  {"adp": "PC34", "name": "PC 4 Flex 4.1",                             "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC035":  {"adp": "PC35", "name": "PC 4 UML Eastman 4.1",                      "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC036":  {"adp": "PC36", "name": "PC 4 Boeing 4.5",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC037":  {"adp": "PC37", "name": "PC 4 Boeing 4.7",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC038":  {"adp": "PC38", "name": "PC 4 GE Global 4.8",                        "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC039":  {"adp": "PC39", "name": "PC 4 Boeing 4.9",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC040":  {"adp": "PC40", "name": "PC 4 Lockheed Martin 4.9",                  "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC041":  {"adp": "PC41", "name": "PC 5 Worcester Polytechnic 5.1",            "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC042":  {"adp": "PC42", "name": "PC 5 Binghamton 5.2A",                      "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC043":  {"adp": "PC43", "name": "PC 5 SysteMECH 5.2B",                       "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC044":  {"adp": "PC44", "name": "PC 5 HP 5.3",                               "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC045":  {"adp": "PC45", "name": "PC 5 GE Global 5.4",                        "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC046":  {"adp": "PC46", "name": "PC 5 Binghamton LMC 5.5",                   "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC047":  {"adp": "PC47", "name": "PC 5 Auburn 5.6",                           "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC048":  {"adp": "PC48", "name": "PC 5 Boeing 5.7",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC049":  {"adp": "PC49", "name": "PC 5 GE Global 5.7",                        "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC050":  {"adp": "PC50", "name": "PC 5 GE Global 5.8",                        "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC051":  {"adp": "PC51", "name": "PC 5 Boeing 5.9",                           "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC052":  {"adp": "PC52", "name": "PC 5 Lockheed Martin 5.10",                 "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "1PC053":  {"adp": "PC53", "name": "PC 0 Auburn University",                    "type_b": "PC",  "type_c": "State or Local",  "rollup": None},
    "1PC054":  {"adp": "PC54", "name": "PC 0 GE Global Research",                   "type_b": "PC",  "type_c": "Industry",        "rollup": None},
    "2ADP001": {"adp": 1,   "name": "CA2 ADP1 ESI Laser",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1001"},
    "2ADP002": {"adp": 2,   "name": "CA2 ADP2 BMNT NSIN",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1002"},
    "2ADP003": {"adp": 3,   "name": "CA2 ADP3 Reliability Ph1",                     "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1003"},
    "2ADP004": {"adp": 4,   "name": "CA2 ADP4 DPiX Flex xray",                      "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1004"},
    "2ADP005": {"adp": 5,   "name": "CA2 ADP5",                                     "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP006": {"adp": 6,   "name": "CA2 ADP6 CCDC AC2 nd SubKs",                  "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1006"},
    "2ADP007": {"adp": 7,   "name": "CA2 ADP7",                                     "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP008": {"adp": 8,   "name": "CA2 ADP8 BMNT Collab",                         "type_b": "ADP-EWD, ADP-Tech, ADP-Tech Subk",       "type_c": "Other DOD",       "rollup": "1008"},
    "2ADP009": {"adp": 9,   "name": "CA2 ADP9 Palo Alto Research Center",           "type_b": "ADP-EWD, ADP-Tech, ADP-Tech Subk, PC ADP","type_c": "Other DOD",      "rollup": "1009"},
    "2ADP010": {"adp": 10,  "name": "CA2 ADP10",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP011": {"adp": 11,  "name": "CA2 ADP11 WFD Project",                        "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Other DOD",       "rollup": "1011"},
    "2ADP012": {"adp": 12,  "name": "CA2 ADP12",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP013": {"adp": 13,  "name": "CA2 ADP13",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP014": {"adp": 14,  "name": "CA2 ADP14",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP015": {"adp": 15,  "name": "CA2 ADP15",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP016": {"adp": 16,  "name": "CA2 ADP16 NF",                                 "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Other DOD",       "rollup": "1016"},
    "2ADP017": {"adp": 17,  "name": "CA2 ADP17",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP018": {"adp": 18,  "name": "CA2 ADP18",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP019": {"adp": 19,  "name": "CA2 ADP19 NASIC BMNT GXF",                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1019"},
    "2ADP020": {"adp": 20,  "name": "CA2 ADP20 Base unassigned PC funds",           "type_b": "ADP-Tech, ADP-Tech Subk, PC ADP",         "type_c": "Airforce",        "rollup": "1020"},
    "2ADP021": {"adp": 21,  "name": "CA2 ADP21",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP022": {"adp": 22,  "name": "CA2 ADP22",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP023": {"adp": 23,  "name": "CA2 ADP23",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1023"},
    "2ADP024": {"adp": 24,  "name": "CA2 ADP24",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP025": {"adp": 25,  "name": "CA2 ADP25 Program Management",                 "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1025"},
    "2ADP026": {"adp": 26,  "name": "CA2 ADP26",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP027": {"adp": 27,  "name": "CA2 ADP27 QMS",                                "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1027"},
    "2ADP028": {"adp": 28,  "name": "CA2 ADP28",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP029": {"adp": 29,  "name": "CA2 ADP29",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP030": {"adp": 30,  "name": "CA2 ADP30 ACI",                                "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1030"},
    "2ADP031": {"adp": 31,  "name": "CA2 ADP31",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP033": {"adp": 33,  "name": "CA2 ADP33 WFD Training",                       "type_b": "ADP-EWD, ADP-Tech, Core, PC ADP",         "type_c": "Other DOD",       "rollup": "1033"},
    "2ADP034": {"adp": 34,  "name": "CA2 ADP34 MaxPower",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1034"},
    "2ADP035": {"adp": 35,  "name": "CA2 ADP35",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1035"},
    "2ADP036": {"adp": 36,  "name": "CA2 ADP36",                                    "type_b": "ADP-Tech",                                "type_c": "Other Gov",       "rollup": None},
    "2ADP037": {"adp": 37,  "name": "CA2 ADP37 NF Program Management",              "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1037"},
    "2ADP038": {"adp": 38,  "name": "CA2 ADP38",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP039": {"adp": 39,  "name": "CA2 ADP39",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP040": {"adp": 40,  "name": "CA2 ADP40",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP041": {"adp": 41,  "name": "CA2 ADP41",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP042": {"adp": 42,  "name": "CA2 ADP42",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP043": {"adp": 43,  "name": "CA2 ADP43",                                    "type_b": "ADP-Tech",                                "type_c": "Other Gov",       "rollup": None},
    "2ADP044": {"adp": 44,  "name": "CA2 ADP44",                                    "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": None},
    "2ADP045": {"adp": 45,  "name": "CA2 ADP45",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP046": {"adp": 46,  "name": "CA2 ADP46 Innov Day",                          "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1046"},
    "2ADP047": {"adp": 47,  "name": "CA2 ADP47",                                    "type_b": "ADP-Tech Subk",                           "type_c": "Army",            "rollup": None},
    "2ADP048": {"adp": 48,  "name": "CA2 ADP48 BMNT",                               "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1048"},
    "2ADP049": {"adp": 49,  "name": "CA2 ADP49",                                    "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Other DOD",       "rollup": "1049"},
    "2ADP050": {"adp": 50,  "name": "CA2 ADP50",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP051": {"adp": 51,  "name": "CA2 ADP51",                                    "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "2ADP052": {"adp": 52,  "name": "CA2 ADP52",                                    "type_b": "ADP-Tech",                                "type_c": "Other Gov",       "rollup": None},
    "2ADP053": {"adp": 53,  "name": "CA2 ADP53",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP054": {"adp": 54,  "name": "CA2 ADP54",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1054"},
    "2ADP055": {"adp": 55,  "name": "CA2 ADP55 SWRI",                               "type_b": "ADP-Tech Subk",                           "type_c": "Airforce",        "rollup": "1055"},
    "2ADP056": {"adp": 56,  "name": "CA2 ADP56",                                    "type_b": "ADP-Tech",                                "type_c": "Army",            "rollup": None},
    "2ADP057": {"adp": 57,  "name": "CA2 ADP57 GE Healthcare",                      "type_b": "ADP-Tech, PC ADP",                        "type_c": "Other DOD",       "rollup": "1057"},
    "2ADP058": {"adp": 58,  "name": "CA2 ADP58 SubK BEST",                          "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Other DOD",       "rollup": "1058"},
    "2ADP059": {"adp": 59,  "name": "CA2 ADP59",                                    "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Army, Other DOD", "rollup": "1059"},
    "2ADP060": {"adp": 60,  "name": "CA2 ADP60 SUBK GE",                            "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1060"},
    "2ADP061": {"adp": 61,  "name": "CA2 ADP61",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1061"},
    "2ADP062": {"adp": 62,  "name": "CA2 ADP62 Lightcast",                          "type_b": "ADP-EWD, ADP-EWD Subk",                  "type_c": "Other DOD",       "rollup": "1062"},
    "2ADP063": {"adp": 63,  "name": "CA2 ADP63 Sciperio",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1063"},
    "2ADP064": {"adp": 64,  "name": "CA2 ADP64 NF PC Support",                      "type_b": "ADP-Tech, PC ADP",                        "type_c": "Other DOD",       "rollup": "1064"},
    "2ADP065": {"adp": 65,  "name": "CA2 ADP65 NF Mgt",                             "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1065"},
    "2ADP066": {"adp": 66,  "name": "CA2 ADP66 NF Eng",                             "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1066"},
    "2ADP067": {"adp": 67,  "name": "CA2 ADP67 AFFOA",                              "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1067"},
    "2ADP068": {"adp": 68,  "name": "CA2 ADP68 LSU",                                "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1068"},
    "2ADP069": {"adp": 69,  "name": "CA2 ADP69",                                    "type_b": "ADP-Tech",                                "type_c": "Other DOD",       "rollup": None},
    "2ADP070": {"adp": 70,  "name": "CA2 ADP70 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1070"},
    "2ADP071": {"adp": 71,  "name": "CA2 ADP71 Boise Support",                      "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1071"},
    "2ADP072": {"adp": 72,  "name": "CA2 ADP72 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other Gov",       "rollup": "1072"},
    "2ADP073": {"adp": 73,  "name": "CA2 ADP73",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1073"},
    "2ADP074": {"adp": 74,  "name": "CA2 ADP74 Semi",                               "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1074"},
    "2ADP075": {"adp": 75,  "name": "CA2 ADP75",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1075"},
    "2ADP076": {"adp": 76,  "name": "CA2 ADP76",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1076"},
    "2ADP077": {"adp": 77,  "name": "CA2 ADP77 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1077"},
    "2ADP078": {"adp": 78,  "name": "CA2 ADP78 Lockheed Martin",                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1078"},
    "2ADP079": {"adp": 79,  "name": "CA2 ADP79 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1079"},
    "2ADP080": {"adp": 80,  "name": "CA2 ADP80 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1080"},
    "2ADP081": {"adp": 81,  "name": "CA2 ADP81 SUBK MGT",                           "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1081"},
    "2ADP082": {"adp": 82,  "name": "CA2 ADP82",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1082"},
    "2ADP083": {"adp": 83,  "name": "CA2 ADP83 Core",                               "type_b": "Core, PC ADP",                            "type_c": "Other DOD",       "rollup": "1083"},
    "2ADP084": {"adp": 84,  "name": "CA2 ADP84 U Mass Lowell",                      "type_b": "ADP-Tech, PC ADP",                        "type_c": "Other DOD",       "rollup": "1084"},
    "2ADP085": {"adp": 85,  "name": "CA2 ADP85",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1085"},
    "2ADP086": {"adp": 86,  "name": "CA2 ADP86 SubK Mgmt",                          "type_b": "ADP-Tech",                                "type_c": "Airforce",        "rollup": "1086"},
    "2ADP087": {"adp": 87,  "name": "CA2 ADP87",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1087"},
    "2ADP088": {"adp": 88,  "name": "CA2 ADP88 NF Subk Mgt",                        "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1088"},
    "2ADP089": {"adp": 89,  "name": "CA2 ADP89 CNM Ingenuity",                      "type_b": "ADP-EWD, ADP-EWD Subk, ADP-Tech",        "type_c": "Other DOD",       "rollup": "1089"},
    "2ADP090": {"adp": 90,  "name": "CA2 ADP90 SubK Graf CIRCA",                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1090"},
    "2ADP091": {"adp": 91,  "name": "CA2 ADP91 SubK Mgmt",                          "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1091"},
    "2ADP092": {"adp": 92,  "name": "CA2 ADP92 SubK Mgmt",                          "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1092"},
    "2ADP093": {"adp": 93,  "name": "CA2 ADP93 Auburn University",                  "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1093"},
    "2ADP094": {"adp": 94,  "name": "CA2 ADP94",                                    "type_b": "ADP-Tech",                                "type_c": "Navy",            "rollup": None},
    "2ADP095": {"adp": 95,  "name": "CA2 ADP95",                                    "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1095"},
    "2ADP096": {"adp": 96,  "name": "CA2 ADP96 SubK Mgmt",                          "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Navy",            "rollup": "1096"},
    "2ADP097": {"adp": 97,  "name": "CA2 ADP97 BMNT",                               "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1097"},
    "2ADP098": {"adp": 98,  "name": "CA2 ADP98 Boeing",                             "type_b": "ADP-EWD, ADP-Tech, PC ADP",               "type_c": "Other DOD",       "rollup": "1098"},
    "2ADP099": {"adp": 99,  "name": "CA2 ADP99",                                    "type_b": None,                                      "type_c": "Other DOD",       "rollup": "1099"},
    "2ADP100": {"adp": 100, "name": "CA2 ADP100 NF Tech Support",                   "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Other DOD",       "rollup": "1100"},
    "2ADP101": {"adp": 101, "name": "CA2 ADP101 NF Support for Comet",              "type_b": "ADP-Tech, Core",                          "type_c": "Other DOD",       "rollup": "1101"},
    "2ADP102": {"adp": 102, "name": "CA2 ADP102",                                   "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1102"},
    "2ADP103": {"adp": 103, "name": "CA2 ADP103",                                   "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Army",            "rollup": "1103"},
    "2ADP104": {"adp": 104, "name": "CA2 ADP104 FAB",                               "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1104"},
    "2ADP106": {"adp": 106, "name": "CA2 ADP106 Subk Mgmt",                         "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1106"},
    "2ADP107": {"adp": 107, "name": "CA2 ADP107 FAB",                               "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Navy",            "rollup": "1107"},
    "2ADP109": {"adp": 109, "name": "CA2 ADP109",                                   "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1109"},
    "2ADP110": {"adp": 110, "name": "CA2 ADP110 SubK Mgmt",                         "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1110"},
    "2ADP112": {"adp": 112, "name": "CA2 ADP112",                                   "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1112"},
    "2ADP113": {"adp": 113, "name": "CA2 ADP113 FAB & ENG",                         "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Airforce",        "rollup": "1113"},
    "2ADP114": {"adp": 114, "name": "CA2 ADP114 SUBK Mgmt",                         "type_b": "ADP-Tech, ADP-Tech Subk",                 "type_c": "Navy",            "rollup": "1114"},
}

# ── Styles ────────────────────────────────────────────────────────────────────
YELLOW_FILL  = PatternFill("solid", fgColor="F4F169")   # Contracting sidebar
PINK_FILL    = PatternFill("solid", fgColor="FFB3C1")   # govt obligated
GREEN_FILL   = PatternFill("solid", fgColor="B7E1CD")   # cash collected
GREY_FILL    = PatternFill("solid", fgColor="F2F2F2")

BOLD  = Font(bold=True)
NORMAL = Font(bold=False)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
VCENTER = Alignment(horizontal="left",  vertical="center", wrap_text=True)

NUM_FMT  = '#,##0.00;(#,##0.00);"-"'
DASH_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─────────────────────────────────────────────────────────────────────────────
# API FETCHING
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_netamount(sequence: str, fye: str, quarter: str,
                     segment: str, account_name: str,
                     retries: int = 3) -> float:
    params = urllib.parse.urlencode({
        "display_sequence": sequence,
        "fiscal_year_end":  fye,
        "fiscal_quarter":   quarter,
        "segment":          segment,
        "account_name":     account_name,
    })
    url = f"{NETAMOUNT_URL}?{params}"
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                data = json.loads(r.read())
            val = data.get("items", [{}])[0].get("total_netamount")
            return float(val) if val is not None else 0.0
        except Exception as exc:
            if attempt == retries - 1:
                logger.warning("FAILED %s %s %s %s: %s", sequence, fye, quarter, account_name, exc)
                return 0.0
            time.sleep(1.5 ** attempt)
    return 0.0


def _build_tasks(sequence: str) -> list:
    tasks = []
    for year in FISCAL_YEARS:
        fye = f"FYE 9/30/{year}"
        for quarter in QUARTERS:
            for _, _, account in ACTUALS_BUDGET_ACCOUNTS:
                if account is not None:
                    tasks.append((fye, quarter, "ACTUALS",     account))
                    tasks.append((fye, quarter, "BUDGETED",    account))
            for _, account in CONTRACTING_ACCOUNTS:
                tasks.append((fye, quarter, "CONTRACTING", account))
    return tasks


def fetch_sequence_data(sequence: str, workers: int = 20) -> dict:
    tasks = _build_tasks(sequence)
    results = {}
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {
            pool.submit(_fetch_netamount, sequence, fye, q, seg, acc): (fye, q, seg, acc)
            for (fye, q, seg, acc) in tasks
        }
        for future in as_completed(future_map):
            key = future_map[future]
            results[key] = future.result()
    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────

def _fye(year: int) -> str:
    return f"FYE 9/30/{year}"


def _quarter_label(year: int, q: str) -> str:
    """e.g. 2020, Q1 -> Q1 FY20"""
    return f"{q} FY{str(year)[2:]}"


def _data_start_col() -> int:
    """Data columns start at col C = 3  (A=section, B=account name)"""
    return 3


def _write_col_headers(ws, fiscal_years: list) -> dict:
    """
    Write FYE headers row 9 and Q1-Q4 row 10.
    Returns col_map: { (fye, quarter) -> col_index }
    """
    col = _data_start_col()
    col_map = {}

    for year in fiscal_years:
        fye = _fye(year)
        # Merge FYE label across 4 Q cols
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 3)
        c = ws.cell(row=9, column=col, value=fye)
        c.alignment = CENTER
        for q in QUARTERS:
            qc = ws.cell(row=10, column=col, value=q)
            qc.alignment = CENTER
            col_map[(fye, q)] = col
            col += 1

    # Total column header
    ws.cell(row=9, column=col, value="Total").alignment = CENTER
    ws.cell(row=10, column=col).alignment = CENTER
    col_map["total"] = col

    return col_map


def _write_data_row(ws, row: int, data: dict, fye_q_pairs: list,
                    col_map: dict, segment: str, api_account: Optional[str]):
    """Write one data row across all FYE/Quarter columns + Total formula."""
    first_col = None
    last_col = None
    for (fye, q) in fye_q_pairs:
        col = col_map.get((fye, q))
        if col is None:
            continue
        if first_col is None:
            first_col = col
        last_col = col
        amount = 0.0 if api_account is None else data.get((fye, q, segment, api_account), 0.0)
        if amount != 0.0:
            c = ws.cell(row=row, column=col, value=amount)
            c.number_format = DASH_FMT
            c.alignment = RIGHT

    # Total column: =SUM(first_data_col:last_data_col)
    total_col = col_map.get("total")
    if total_col and first_col and last_col:
        fl = get_column_letter(first_col)
        ll = get_column_letter(last_col)
        tc = ws.cell(row=row, column=total_col,
                     value=f"=SUM({fl}{row}:{ll}{row})")
        tc.number_format = DASH_FMT
        tc.alignment = RIGHT


def _write_total_row(ws, row: int, data_start_row: int, col_map: dict):
    """Write a SUM total row (sums account rows, skipping Labor Hours)."""
    ws.cell(row=row, column=2, value="Total").font = BOLD
    # SUM from data_start_row+1 to skip Labor Hours (row 11/25)
    sum_start = data_start_row + 1
    for col_idx in col_map.values():
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=row, column=col_idx,
                    value=f"=SUM({col_letter}{sum_start}:{col_letter}{row - 1})")
        c.number_format = DASH_FMT
        c.alignment = RIGHT


def _write_formulas_section(ws, fiscal_years: list, col_map: dict,
                             actuals_total_row: int, budgeted_total_row: int,
                             committed_row: int, obligated_row: int,
                             expended_row: int):
    """
    Write the cumulative running-total formulas section starting at row 48.

    Template pattern (from FEMR_for_NetSuite.xlsx ADP 33):
      First col:  =C39            (direct reference to source row)
      Next cols:  =D39+C49        (current quarter + previous cumulative)

    This produces a running total across all quarters.
    """
    header_row = 48
    ws.cell(row=header_row, column=1, value="formulas")

    # Write quarter labels as "Q1 FY20" etc
    col = _data_start_col()
    for year in fiscal_years:
        for q in QUARTERS:
            ws.cell(row=header_row, column=col, value=_quarter_label(year, q)).alignment = CENTER
            col += 1

    # Formula rows — cumulative running totals referencing source rows
    formula_defs = [
        (49, "Total Committed",  committed_row),
        (50, "Total Obligated",  obligated_row),
        (51, "Total Expended",   expended_row),
        (52, "Budgeted Plan",    budgeted_total_row),
        (53, "Revised Plan",     None),   # no data yet
        (54, "Actual",           actuals_total_row),
    ]

    for (frow, label, src_row) in formula_defs:
        ws.cell(row=frow, column=2, value=label)
        if src_row is None:
            continue

        prev_col_letter = None
        col = _data_start_col()
        for year in fiscal_years:
            for q in QUARTERS:
                col_letter = get_column_letter(col)
                if prev_col_letter is None:
                    # First quarter: direct reference
                    formula = f"={col_letter}{src_row}"
                else:
                    # Subsequent quarters: current + previous cumulative
                    formula = f"={col_letter}{src_row}+{prev_col_letter}{frow}"
                c = ws.cell(row=frow, column=col, value=formula)
                c.number_format = DASH_FMT
                c.alignment = RIGHT
                prev_col_letter = col_letter
                col += 1


def build_adp_sheet(wb: Workbook, sequence: str, meta: dict,
                    data: dict, fiscal_years: list):
    """Create one ADP tab matching the template exactly."""

    # ── Sheet name ────────────────────────────────────────────────────────
    base = f"ADP {meta['adp']}"
    sheet_name = base
    counter = 2
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base}({counter})"
        counter += 1
    ws = wb.create_sheet(title=sheet_name[:31])

    # ── Rows 1–8: Header block (8 rows matching template) ────────────────
    header_fields = [
        ("NextFlex",            ""),
        ("ADP",                 meta["adp"]),
        ("Project Name",        meta["name"]),
        ("External Approver 1", ""),
        ("Project Type",        meta.get("type_b") or ""),
        ("Service",             meta.get("type_c") or ""),
        ("Color of Money",      ""),
        ("ACRNs",               meta.get("rollup") or ""),
    ]
    for r, (label, value) in enumerate(header_fields, start=1):
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2, value=value)
    # Rollup label in col C row 2 if applicable
    if meta.get("rollup"):
        ws.cell(row=2, column=3, value="Rollup")

    # ── Rows 9–10: Column headers ─────────────────────────────────────────
    col_map = _write_col_headers(ws, fiscal_years)
    fye_q_pairs = [(_fye(y), q) for y in fiscal_years for q in QUARTERS]

    # ── Rows 11–24: Actuals section ───────────────────────────────────────
    # Merge A11:A24 for "Actuals" sidebar label
    ws.merge_cells(start_row=11, start_column=1, end_row=24, end_column=1)
    ac = ws.cell(row=11, column=1, value="Actuals")
    ac.font = BOLD
    ac.alignment = VCENTER

    for i, (label, code, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = 11 + i  # rows 11–23
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "ACTUALS", api_account)

    # Row 24: Total (SUM skips Labor Hours at row 11, sums 12–23)
    _write_total_row(ws, 24, 11, col_map)
    actuals_total_row = 24

    # ── Rows 25–38: Budgeted section ──────────────────────────────────────
    ws.merge_cells(start_row=25, start_column=1, end_row=38, end_column=1)
    bc = ws.cell(row=25, column=1, value="Budgeted")
    bc.font = BOLD
    bc.alignment = VCENTER

    for i, (label, code, api_account) in enumerate(ACTUALS_BUDGET_ACCOUNTS):
        row = 25 + i  # rows 25–37
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "BUDGETED", api_account)

    # Row 38: Total (SUM skips Labor Hours at row 25, sums 26–37)
    _write_total_row(ws, 38, 25, col_map)
    budgeted_total_row = 38

    # ── Rows 39–41: Contracting section ──────────────────────────────────
    ws.merge_cells(start_row=39, start_column=1, end_row=41, end_column=1)
    cc = ws.cell(row=39, column=1, value="Contracting")
    cc.alignment = VCENTER
    # Yellow fill for contracting label
    for r in range(39, 42):
        ws.cell(row=r, column=1).fill = YELLOW_FILL

    committed_row  = 39
    obligated_row  = 40
    expended_row   = 41

    for i, (label, api_account) in enumerate(CONTRACTING_ACCOUNTS):
        row = 39 + i
        ws.cell(row=row, column=2, value=label)
        _write_data_row(ws, row, data, fye_q_pairs, col_map, "CONTRACTING", api_account)

    # ── Row 42: blank ─────────────────────────────────────────────────────

    # ── Rows 43–46: Govt awards calc fields ──────────────────────────────
    ws.cell(row=43, column=2, value="govt awards")
    ws.cell(row=44, column=2, value="govt obligated")
    ws.cell(row=45, column=2, value="cash collected")
    ws.cell(row=46, column=1, value="add calc field")
    ws.cell(row=46, column=2, value="remaining cash")

    # ── Row 47: blank ─────────────────────────────────────────────────────

    # ── Rows 48–54: Formulas summary section (cumulative running totals) ─
    _write_formulas_section(
        ws, fiscal_years, col_map,
        actuals_total_row, budgeted_total_row,
        committed_row, obligated_row, expended_row,
    )

    # ── Source notes rows 59–61 ───────────────────────────────────────────
    ws.cell(row=59, column=1, value="source: bible")
    ws.cell(row=60, column=1, value="source: project bud vs actual")
    ws.cell(row=61, column=1, value="source: SF270 needs to be put NSAW")

    # ── Freeze panes & column widths ─────────────────────────────────────
    ws.freeze_panes = "C11"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    for year in fiscal_years:
        for q in QUARTERS:
            col = col_map.get((_fye(year), q))
            if col:
                ws.column_dimensions[get_column_letter(col)].width = 13
    total_col = col_map.get("total")
    if total_col:
        ws.column_dimensions[get_column_letter(total_col)].width = 16

    logger.info("  tab '%s'  (%s)", sheet_name, meta["name"])


# ─────────────────────────────────────────────────────────────────────────────
# FILE SPLITTING
# ─────────────────────────────────────────────────────────────────────────────

def _output_filename(base: str, adp_start: int, adp_end: int) -> str:
    """
    Generate output filename for a range chunk.
    e.g. femr_ca2_report.xlsx + (1, 50) -> femr_ca2_report_adp001-050.xlsx
    """
    stem = base.replace(".xlsx", "")
    return f"{stem}_adp{adp_start:03d}-{adp_end:03d}.xlsx"


def _chunk_sequences_by_adp_range(sequences: list, chunk_size: int = 50) -> list:
    """
    Split sequences into chunks based on ADP number ranges.
    Each chunk covers a range of `chunk_size` ADP numbers (e.g. 1-50, 51-100).
    Sequences with non-integer ADP numbers (e.g. PC1, PC2) go into a separate chunk.

    Returns: list of (adp_range_start, adp_range_end, [sequences_in_range])
    """
    numeric = [(seq, meta) for seq, meta in
               ((s, ADP_REGISTRY[s]) for s in sequences)
               if isinstance(meta["adp"], int)]
    non_numeric = [(seq, meta) for seq, meta in
                   ((s, ADP_REGISTRY[s]) for s in sequences)
                   if not isinstance(meta["adp"], int)]

    if not numeric:
        return []

    max_adp = max(meta["adp"] for _, meta in numeric)

    chunks = []
    range_start = 1
    while range_start <= max_adp:
        range_end = range_start + chunk_size - 1
        chunk_seqs = [seq for seq, meta in numeric
                      if range_start <= meta["adp"] <= range_end]
        if chunk_seqs:
            chunks.append((range_start, range_end, chunk_seqs))
        range_start = range_end + 1

    if non_numeric:
        chunks.append(("PC", "PC", [seq for seq, _ in non_numeric]))

    return chunks


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(output_file: str = "femr_netsuite_report.xlsx",
        filter_sequence: Optional[str] = None,
        ca2_only: bool = False,
        workers: int = 20,
        chunk_size: int = 50):

    logger.info("=== FEMR NetSuite Report Generator (v4) ===")
    logger.info("Fiscal years: %s", FISCAL_YEARS)

    sequences = sorted(ADP_REGISTRY.keys())
    if filter_sequence:
        sequences = [s for s in sequences if s == filter_sequence]
        if not sequences:
            logger.error("Sequence '%s' not found.", filter_sequence)
            sys.exit(1)
        # Single sequence — just write one file directly
        meta = ADP_REGISTRY[sequences[0]]
        wb = Workbook()
        wb.remove(wb.active)
        data = fetch_sequence_data(sequences[0], workers=workers)
        build_adp_sheet(wb, sequences[0], meta, data, FISCAL_YEARS)
        wb.save(output_file)
        logger.info("Done! Saved to %s", output_file)
        return
    elif ca2_only:
        sequences = [s for s in sequences if s.startswith("2ADP")]

    total = len(sequences)
    tasks_per = len(_build_tasks("_"))
    logger.info("Sequences  : %d", total)
    logger.info("API calls  : ~%d", total * tasks_per)
    logger.info("Workers    : %d", workers)
    logger.info("Chunk size : %d ADPs per file", chunk_size)

    # Split into range-based chunks
    chunks = _chunk_sequences_by_adp_range(sequences, chunk_size)
    logger.info("Output files: %d", len(chunks))
    for adp_start, adp_end, chunk_seqs in chunks:
        fname = _output_filename(output_file, adp_start, adp_end) \
                if isinstance(adp_start, int) else output_file.replace(".xlsx", "_pc.xlsx")
        logger.info("  %s  (%d tabs)", fname, len(chunk_seqs))

    # Process each chunk
    overall = 0
    for adp_start, adp_end, chunk_seqs in chunks:
        if isinstance(adp_start, int):
            fname = _output_filename(output_file, adp_start, adp_end)
        else:
            fname = output_file.replace(".xlsx", "_pc.xlsx")

        logger.info("\n--- File: %s (%d tabs) ---", fname, len(chunk_seqs))

        wb = Workbook()
        wb.remove(wb.active)

        for i, seq in enumerate(chunk_seqs, 1):
            overall += 1
            meta = ADP_REGISTRY[seq]
            logger.info("[%d/%d overall | %d/%d in file] %s — %s",
                        overall, total, i, len(chunk_seqs), seq, meta["name"])
            data = fetch_sequence_data(seq, workers=workers)
            build_adp_sheet(wb, seq, meta, data, FISCAL_YEARS)

            # Checkpoint every 10 tabs
            if i % 10 == 0:
                wb.save(fname)
                logger.info("  Checkpoint saved (%d/%d in file)", i, len(chunk_seqs))

        wb.save(fname)
        logger.info("Saved: %s  (%d tabs)", fname, len(chunk_seqs))

    logger.info("\nAll done! %d tabs across %d files.", total, len(chunks))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate FEMR NetSuite multi-tab Excel report (v4), split into files of 50 ADPs."
    )
    parser.add_argument("--output",     "-o", default="femr_netsuite_report.xlsx",
                        help="Base output filename (range suffix added automatically)")
    parser.add_argument("--sequence",   "-s", default=None,
                        help="Single sequence only e.g. 2ADP001 (for testing)")
    parser.add_argument("--ca2-only",   action="store_true",
                        help="Only generate CA2 ADP tabs")
    parser.add_argument("--workers",    "-w", type=int, default=20,
                        help="Concurrent threads per sequence (default 20)")
    parser.add_argument("--chunk-size", "-c", type=int, default=50,
                        help="Max ADP range per output file (default 50)")
    args = parser.parse_args()
    run(output_file=args.output,
        filter_sequence=args.sequence,
        ca2_only=args.ca2_only,
        workers=args.workers,
        chunk_size=args.chunk_size)
