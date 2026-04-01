"""
FEMR Funds Transformation Script
Reads 'FEMR Funds' input sheet and generates 'Output' tab in the same format.

Output structure:
  For each quarter date (26 quarters), 4 blocks of 134 rows each:
    Block 1: Committed (Award Amount)
    Block 2: Obligated
    Block 3: Expended (Cash Collected)
    Block 4: Remaining Cash (= Obligated - Expended for that quarter)

Column mapping in FEMR Funds (1-indexed):
  Col 4: Sequence
  Col 10-12: FY20 Q3 (6/30/20)
  Col 13-15: FY20 Q4 (9/30/20)
  [skip 16-18: FY20 Total]
  Col 19-21: FY21 Q1 (12/31/20)
  Col 22-24: FY21 Q2 (3/30/21)
  Col 25-27: FY21 Q3 (6/30/21)
  Col 28-30: FY21 Q4 (9/30/21)
  [skip 31-33: FY21 Total]
  ... and so on, skipping annual totals
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import copy
import argparse

# ─── Quarter definitions: (date_str, col_committed, col_obligated, col_expended)
# Columns are 1-indexed matching the FEMR Funds sheet
QUARTERS = [
    (datetime(2020,  6, 30), 10, 11, 12),
    (datetime(2020,  9, 30), 13, 14, 15),
    # skip FY20 Total (16-18)
    (datetime(2020, 12, 31), 19, 20, 21),
    (datetime(2021,  3, 30), 22, 23, 24),
    (datetime(2021,  6, 30), 25, 26, 27),
    (datetime(2021,  9, 30), 28, 29, 30),
    # skip FY21 Total (31-33)
    (datetime(2021, 12, 31), 34, 35, 36),
    (datetime(2022,  3, 30), 37, 38, 39),
    (datetime(2022,  6, 30), 40, 41, 42),
    (datetime(2022,  9, 30), 43, 44, 45),
    # skip FY22 Total (46-48)
    (datetime(2022, 12, 31), 49, 50, 51),
    (datetime(2023,  3, 30), 52, 53, 54),
    (datetime(2023,  6, 30), 55, 56, 57),
    (datetime(2023,  9, 30), 58, 59, 60),
    # skip FY23 Total (61-63)
    (datetime(2023, 12, 31), 64, 65, 66),
    (datetime(2024,  3, 30), 67, 68, 69),
    (datetime(2024,  6, 30), 70, 71, 72),
    (datetime(2024,  9, 30), 73, 74, 75),
    # skip FY24 Total (76-78)
    (datetime(2024, 12, 31), 79, 80, 81),
    (datetime(2025,  3, 30), 82, 83, 84),
    (datetime(2025,  6, 30), 85, 86, 87),
    (datetime(2025,  9, 30), 88, 89, 90),
    # skip FY25 Total (91-93)
    (datetime(2025, 12, 31), 94, 95, 96),
    (datetime(2026,  3, 30), 97, 98, 99),
    (datetime(2026,  6, 30), 100, 101, 102),
    (datetime(2026,  9, 30), 103, 104, 105),
    # skip FY26 Total (106-108), Total (109-112)
]

# Fallback paths (the script is now meant to be called with CLI args)
INPUT_FILE = '/mnt/user-data/uploads/2026_03_FEMR_funds.xlsx'
OUTPUT_FILE = '/mnt/user-data/outputs/2026_03_FEMR_funds_output.xlsx'

def safe_float(val):
    """Convert value to float, returning 0.0 if None or not numeric."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0

def read_input(wb):
    """
    Read FEMR Funds sheet and aggregate values by (sequence, quarter).
    Returns:
        sequences: ordered list of unique sequences (from Output tab)
        data: dict { sequence -> { quarter_date -> (committed, obligated, expended) } }
    """
    ws = wb['FEMR Funds']

    # Sequence order is taken from the existing Output template when present.
    if 'Output' in wb.sheetnames:
        ws_out = wb['Output']
        sequences = []
        for row in ws_out.iter_rows(min_row=2, max_row=135, values_only=True):
            if row[0] is not None:
                sequences.append(row[0])
    else:
        # Fallback: derive sequences from the FEMR Funds sheet, preserving first-seen order.
        sequences = []
        for row in ws.iter_rows(min_row=7, max_row=306, values_only=True):
            raw_seq = row[3]  # Column D
            if raw_seq is None:
                continue
            seq = str(raw_seq).strip()
            if seq and seq not in sequences:
                sequences.append(seq)

    # Aggregate input data: sequence -> quarter_date -> [committed, obligated, expended]
    data = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0, 0.0]))

    for row in ws.iter_rows(min_row=7, max_row=306, values_only=True):
        raw_seq = row[3]  # Column D (1-indexed col 4, 0-indexed col 3)
        if raw_seq is None:
            continue
        seq = str(raw_seq).strip()
        if seq not in sequences:
            continue

        for (qdate, cc, co, ce) in QUARTERS:
            # cols are 1-indexed; row tuple is 0-indexed
            committed = safe_float(row[cc - 1])
            obligated = safe_float(row[co - 1])
            expended  = safe_float(row[ce - 1])
            data[seq][qdate][0] += committed
            data[seq][qdate][1] += obligated
            data[seq][qdate][2] += expended

    return sequences, data

def generate_output(sequences, data):
    """
    Build the output as a list of rows:
    [sequence, qtr_date, type, amount]
    Structure: for each quarter, 4 blocks (Committed, Obligated, Expended, Remaining Cash),
               each block listing all sequences in order.
    """
    rows = []
    for (qdate, cc, co, ce) in QUARTERS:
        for type_name, idx in [('Committed', 0), ('Obligated', 1), ('Expended', 2)]:
            for seq in sequences:
                amount = data[seq][qdate][idx]
                rows.append((seq, qdate, type_name, amount))
        # Remaining Cash block
        for seq in sequences:
            obligated = data[seq][qdate][1]
            expended  = data[seq][qdate][2]
            remaining = obligated - expended
            rows.append((seq, qdate, 'Remaining Cash', remaining))
    return rows

def copy_sheet_with_formatting(src_wb, src_sheet_name, dst_wb, dst_sheet_name):
    """Copy a sheet with its data and basic formatting to another workbook."""
    src_ws = src_wb[src_sheet_name]
    if dst_sheet_name in dst_wb.sheetnames:
        del dst_wb[dst_sheet_name]
    dst_ws = dst_wb.create_sheet(dst_sheet_name)

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font = copy.copy(cell.font)
                dst_cell.border = copy.copy(cell.border)
                dst_cell.fill = copy.copy(cell.fill)
                dst_cell.number_format = cell.number_format
                dst_cell.alignment = copy.copy(cell.alignment)

    for col in src_ws.column_dimensions:
        dst_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width
    for row in src_ws.row_dimensions:
        dst_ws.row_dimensions[row].height = src_ws.row_dimensions[row].height
    return dst_ws

def write_output_sheet(wb, sequences, data):
    """Write the Output sheet with calculated data."""
    if 'Output' in wb.sheetnames:
        ws = wb['Output']
    else:
        ws = wb.create_sheet('Output')

    output_rows = generate_output(sequences, data)

    # Clear existing data rows (keep header row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write header
    ws['A1'] = 'Sequence'
    ws['B1'] = 'Qtr Date'
    ws['C1'] = 'Type'
    ws['D1'] = 'Amount'

    # Style header
    header_font = Font(bold=True)
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].font = header_font

    # Write data rows
    date_fmt = 'MM/DD/YYYY'
    number_fmt = '#,##0.00;(#,##0.00);-'

    for i, (seq, qdate, type_name, amount) in enumerate(output_rows, start=2):
        ws.cell(row=i, column=1, value=seq)
        date_cell = ws.cell(row=i, column=2, value=qdate)
        date_cell.number_format = date_fmt
        ws.cell(row=i, column=3, value=type_name)
        amount_cell = ws.cell(row=i, column=4, value=round(amount, 2) if amount != 0 else 0)
        amount_cell.number_format = number_fmt

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    print(f"Output sheet written: {len(output_rows)} data rows")

def main(input_file: str = INPUT_FILE, output_file: str = OUTPUT_FILE):
    print("Reading input file...")
    wb = load_workbook(input_file, data_only=True)

    print("Parsing FEMR Funds input...")
    sequences, data = read_input(wb)
    print(f"  Sequences: {len(sequences)}")
    print(f"  Quarters:  {len(QUARTERS)}")
    print(f"  Expected output rows: {len(sequences) * len(QUARTERS) * 4}")

    # Verify a known value: 2ADP001, 9/30/20 should be 27100000 committed
    q = datetime(2020, 9, 30)
    val = data.get('2ADP001', {}).get(q, [0,0,0])
    print(f"\nVerification - 2ADP001 at 9/30/20:")
    print(f"  Committed={val[0]:,.0f}  Obligated={val[1]:,.0f}  Expended={val[2]:,.0f}")

    # 2ADP011 combined: 3175900 + (-837974) = 2337926
    val11 = data.get('2ADP011', {}).get(datetime(2020, 9, 30), [0,0,0])
    print(f"\nVerification - 2ADP011 at 9/30/20:")
    print(f"  Committed={val11[0]:,.0f}  (expected 2,337,926)")

    print("\nWriting output file...")
    out_wb = load_workbook(input_file)

    if 'Output' in out_wb.sheetnames:
        out_ws = out_wb['Output']
        # Delete all columns beyond D to remove extra template content (preserves A-D styles)
        if out_ws.max_column > 4:
            out_ws.delete_cols(5, out_ws.max_column - 4)
    else:
        out_ws = out_wb.create_sheet('Output', 1)
        headers = ['Sequence', 'Qtr Date', 'Type', 'Amount']
        for col, h in enumerate(headers, 1):
            out_ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    output_rows = generate_output(sequences, data)

    for i, (seq, qdate, type_name, amount) in enumerate(output_rows, start=2):
        out_ws.cell(row=i, column=1, value=seq)
        date_cell = out_ws.cell(row=i, column=2, value=qdate)
        date_cell.number_format = 'mm-dd-yy'
        out_ws.cell(row=i, column=3, value=type_name)
        out_ws.cell(row=i, column=4, value=amount)

    # Auto-filter covering full data range (matches original)
    last_row = 1 + len(output_rows)
    out_ws.auto_filter.ref = f'A1:D{last_row}'

    out_wb.save(output_file)
    print(f"Saved: {output_file}")
    print(f"Total data rows written: {len(output_rows)}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="FEMR Funds -> Output transformer")
    parser.add_argument(
        "--input",
        dest="input_file",
        default=INPUT_FILE,
        help="Path to input .xlsx (must include sheet `FEMR Funds` and template `Output` for sequence ordering).",
    )
    parser.add_argument(
        "--output",
        dest="output_file",
        default=OUTPUT_FILE,
        help="Path for the generated output .xlsx.",
    )
    args = parser.parse_args()
    main(args.input_file, args.output_file)
