"""
FEMR Funds Transformation Script v2
Reads 'FEMR Funds' input sheet and generates transformed output data.

Changes from v1:
- Excel output contains ONLY the Output tab (no input tabs carried over)
- Added --format parameter: 'excel' (default) or 'csv'
- Output filename auto-derived from input: output_<input_filename>.<ext>
  (--output still accepted as an explicit override)

Output structure:
  For each quarter date, 4 blocks of N rows each:
    Block 1: Committed (Award Amount)
    Block 2: Obligated
    Block 3: Expended (Cash Collected)
    Block 4: Remaining Cash (= Obligated - Expended for that quarter)

Quarters are discovered dynamically from row 4 of the FEMR Funds sheet.
Row 4 contains labels like 'QE 6/30/20' for quarter end dates and
'FY20 Total' / 'Total' for annual totals (which are automatically skipped).
Each quarter at column C maps to: C=Committed, C+1=Obligated, C+2=Expended.
"""

import csv
import os
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from datetime import datetime
from collections import defaultdict


def discover_quarters(ws):
    """
    Dynamically build the quarter list by reading row 4 of the FEMR Funds sheet.
    Cells in row 4 with a 'QE M/D/YY' label are quarter end dates; all other
    labels (e.g. 'FY20 Total', 'Total') are skipped automatically.

    Returns: list of (quarter_end_date, col_committed, col_obligated, col_expended)
             Columns are 1-indexed.
    """
    quarters = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=4, column=col).value
        if raw is None:
            continue
        label = str(raw).strip()
        if not label.upper().startswith('QE '):
            continue
        date_str = label[3:].strip()
        try:
            qdate = datetime.strptime(date_str, '%m/%d/%y')
        except ValueError:
            continue
        quarters.append((qdate, col, col + 1, col + 2))
    return quarters


def safe_float(val):
    """Convert value to float, returning 0.0 if None or not numeric."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def read_input(wb):
    """
    Read FEMR Funds sheet and aggregate values by (sequence, quarter).
    Quarters are discovered dynamically from row 4 of the FEMR Funds sheet.
    Returns:
        quarters:  list of (quarter_end_date, col_committed, col_obligated, col_expended)
        sequences: ordered list of unique sequences (from Output tab if present)
        data:      dict { sequence -> { quarter_date -> [committed, obligated, expended] } }
    """
    ws = wb['FEMR Funds']
    quarters = discover_quarters(ws)

    # Always derive sequences from FEMR Funds (source of truth).
    sequences = []
    seen = set()
    for row in ws.iter_rows(min_row=7, max_row=306, values_only=True):
        raw_seq = row[3]  # Column D
        if raw_seq is None:
            continue
        seq = str(raw_seq).strip()
        if seq and seq not in seen:
            sequences.append(seq)
            seen.add(seq)

    # If Output template exists in the input file, use its sequence order as a guide.
    # Sequences present in the template are sorted to the front in template order;
    # any new sequences not in the template are appended at the end.
    if 'Output' in wb.sheetnames:
        ws_out = wb['Output']
        template_order = {}
        for row in ws_out.iter_rows(min_row=2, max_row=135, values_only=True):
            if row[0] is not None:
                seq = str(row[0]).strip()
                if seq and seq not in template_order:
                    template_order[seq] = len(template_order)
        if template_order:
            sequences.sort(key=lambda s: template_order.get(s, len(template_order)))

    # Aggregate input data: sequence -> quarter_date -> [committed, obligated, expended]
    data = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0, 0.0]))

    for row in ws.iter_rows(min_row=7, max_row=306, values_only=True):
        raw_seq = row[3]  # Column D (1-indexed col 4, 0-indexed col 3)
        if raw_seq is None:
            continue
        seq = str(raw_seq).strip()
        if seq not in sequences:
            continue

        for (qdate, cc, co, ce) in quarters:
            data[seq][qdate][0] += safe_float(row[cc - 1])
            data[seq][qdate][1] += safe_float(row[co - 1])
            data[seq][qdate][2] += safe_float(row[ce - 1])

    return quarters, sequences, data


def generate_output(quarters, sequences, data):
    """
    Build the output as a list of rows: (sequence, qtr_date, type, amount)
    Structure: for each quarter, 4 blocks (Committed, Obligated, Expended, Remaining Cash),
               each block listing all sequences in order.
    """
    rows = []
    for (qdate, *_) in quarters:
        for type_name, idx in [('Committed', 0), ('Obligated', 1), ('Expended', 2)]:
            for seq in sequences:
                rows.append((seq, qdate, type_name, data[seq][qdate][idx]))
        for seq in sequences:
            remaining = data[seq][qdate][1] - data[seq][qdate][2]
            rows.append((seq, qdate, 'Remaining Cash', remaining))
    return rows


def derive_output_path(input_file, fmt):
    """
    Auto-derive output path from input filename.
    Places output in the same directory as input.
    Prefixes filename with 'output_' and changes extension to match format.
    """
    directory = os.path.dirname(os.path.abspath(input_file))
    basename = os.path.basename(input_file)
    name_no_ext = os.path.splitext(basename)[0]
    ext = '.xlsx' if fmt == 'excel' else '.csv'
    return os.path.join(directory, f'output_{name_no_ext}{ext}')


def write_excel(output_rows, output_file):
    """Write output data to a new Excel workbook with only the Output tab."""
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Output'

    # Header row
    headers = ['Sequence', 'Qtr Date', 'Type', 'Amount']
    for col, h in enumerate(headers, 1):
        out_ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    # Data rows
    for i, (seq, qdate, type_name, amount) in enumerate(output_rows, start=2):
        out_ws.cell(row=i, column=1, value=seq)
        date_cell = out_ws.cell(row=i, column=2, value=qdate)
        date_cell.number_format = 'mm-dd-yy'
        out_ws.cell(row=i, column=3, value=type_name)
        out_ws.cell(row=i, column=4, value=amount)

    # Auto-filter
    last_row = 1 + len(output_rows)
    out_ws.auto_filter.ref = f'A1:D{last_row}'

    # Column widths
    out_ws.column_dimensions['A'].width = 12
    out_ws.column_dimensions['B'].width = 14
    out_ws.column_dimensions['C'].width = 16
    out_ws.column_dimensions['D'].width = 16

    out_wb.save(output_file)


def write_csv(output_rows, output_file):
    """Write output data to a CSV file."""
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Sequence', 'Qtr Date', 'Type', 'Amount'])
        for seq, qdate, type_name, amount in output_rows:
            writer.writerow([seq, qdate.strftime('%m/%d/%Y'), type_name, round(amount, 2)])


def main(input_file, output_file, fmt):
    print("Reading input file...")
    wb = load_workbook(input_file, data_only=True)

    print("Parsing FEMR Funds input...")
    quarters, sequences, data = read_input(wb)
    print(f"  Quarters discovered: {len(quarters)}")
    for qdate, cc, *_ in quarters:
        print(f"    {qdate.strftime('%m/%d/%Y')}  (col {cc})")
    print(f"  Sequences: {len(sequences)}")
    print(f"  Expected output rows: {len(sequences) * len(quarters) * 4}")

    # Spot-check verification
    q = datetime(2020, 9, 30)
    val = data.get('2ADP001', {}).get(q, [0, 0, 0])
    print(f"\nVerification - 2ADP001 at 9/30/20:")
    print(f"  Committed={val[0]:,.0f}  Obligated={val[1]:,.0f}  Expended={val[2]:,.0f}")

    val11 = data.get('2ADP011', {}).get(datetime(2020, 9, 30), [0, 0, 0])
    print(f"\nVerification - 2ADP011 at 9/30/20:")
    print(f"  Committed={val11[0]:,.0f}  (expected 2,337,926)")

    print(f"\nWriting output file ({fmt})...")
    output_rows = generate_output(quarters, sequences, data)

    if fmt == 'excel':
        write_excel(output_rows, output_file)
    else:
        write_csv(output_rows, output_file)

    print(f"Saved: {output_file}")
    print(f"Total data rows written: {len(output_rows)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="FEMR Funds -> Output transformer v2")
    parser.add_argument(
        "--input",
        dest="input_file",
        required=True,
        help="Path to input .xlsx (must include sheet 'FEMR Funds').",
    )
    parser.add_argument(
        "--output",
        dest="output_file",
        default=None,
        help="Path for the generated output file. Auto-derived from input if omitted.",
    )
    parser.add_argument(
        "--format",
        dest="fmt",
        choices=['excel', 'csv'],
        default='excel',
        help="Output format: 'excel' (default) or 'csv'.",
    )
    args = parser.parse_args()

    output_file = args.output_file or derive_output_path(args.input_file, args.fmt)
    main(args.input_file, output_file, args.fmt)
